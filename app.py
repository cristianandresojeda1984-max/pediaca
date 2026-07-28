"""
PediAcá — Aplicación Principal
Stack: Python + Flask + SQLite
"""

import os
try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    psycopg2 = None
import re
import io
import random
from functools import wraps
from flask import (Flask, render_template, request, redirect,
                   url_for, session, flash, jsonify, g)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from flask_wtf.csrf import CSRFProtect, CSRFError


# ── PUSH NOTIFICATIONS ────────────────────────────────────────────────────────
import base64, json as _json
from pywebpush import webpush, WebPushException

VAPID_PUBLIC_KEY  = os.environ.get("VAPID_PUBLIC_KEY",  "BDnnssm4HNQJ3uY_EDUaeLr10RVDD9rANxg1Z13OnGL-PAkYM6K_OMf_30aznQofdLphWH2ab5TIwwepauKkd_I")
VAPID_PRIVATE_KEY = os.environ.get("VAPID_PRIVATE_KEY", "LS0tLS1CRUdJTiBFQyBQUklWQVRFIEtFWS0tLS0tCk1IY0NBUUVFSUJWUldGbzY1dXhGSS9mdEk0aTVqdzBIWXZ2d2NEY0pqakRGVHVjejRFSGxvQW9HQ0NxR1NNNDkKQXdFSG9VUURRZ0FFT2VleXliZ2MxQW5lNWo4UU5ScDR1dlhSRlVNUDJzQTNHRFZuWGM2Y1l2NDhDUmd6b3I4NAp4Ly9mUnJPZENoOTB1bUZZZlpwdmxNakRCNmxxNHFSMzhnPT0KLS0tLS1FTkQgRUMgUFJJVkFURSBLRVktLS0tLQo")
VAPID_EMAIL       = os.environ.get("VAPID_CLAIMS_EMAIL", "admin@pediaca.ar")


# ── CONFIGURACIÓN ─────────────────────────────────────────────────────────────
app = Flask(__name__)

app.secret_key = os.environ.get("SECRET_KEY", "dev-key-cambiar-en-produccion")

DB_PATH       = os.environ.get("DB_PATH", "pediaca.db")
DATABASE_URL  = os.environ.get("DATABASE_URL", "")
USE_POSTGRES  = bool(DATABASE_URL and psycopg2)
UPLOAD_FOLDER = "static/uploads"
ALLOWED_EXT   = {"png", "jpg", "jpeg", "webp"}

if app.secret_key == "dev-key-cambiar-en-produccion" and USE_POSTGRES:
    # Sólo advertir (no frenar el arranque): en local con SQLite es normal
    # no tener SECRET_KEY seteada, pero en producción con Postgres sí
    # debería venir de la variable de entorno.
    print("⚠️  SECRET_KEY no configurada — usando la clave de desarrollo. "
          "Configurá SECRET_KEY como variable de entorno en producción.")

# ── CSRF ──────────────────────────────────────────────────────────────────────
# Protege todos los POST/PUT/PATCH/DELETE de la app. base.html se encarga de
# mandar el token en cada <form> (lo inyecta solo por JS) y en cada fetch()
# (header X-CSRFToken), así que no hace falta tocar cada template a mano.
csrf = CSRFProtect(app)


@app.errorhandler(CSRFError)
def _csrf_error(e):
    # Los endpoints llamados por fetch()/JS esperan JSON, no un redirect.
    quiere_json = (
        request.path.startswith("/api/")
        or request.accept_mimetypes.best == "application/json"
        or request.headers.get("Content-Type", "").startswith("application/json")
    )
    if quiere_json:
        return jsonify({"error": "Token de seguridad vencido, recargá la página."}), 400
    flash("La página se venció, probá de nuevo.", "warning")
    return redirect(request.referrer or url_for("home"))

# ── CLOUDINARY (fotos persistentes gratis) ────────────────────────────────────
import cloudinary
import cloudinary.uploader
import cloudinary.api

CLOUDINARY_CLOUD = os.environ.get("CLOUDINARY_CLOUD_NAME", "")
CLOUDINARY_KEY   = os.environ.get("CLOUDINARY_API_KEY",    "")
CLOUDINARY_SEC   = os.environ.get("CLOUDINARY_API_SECRET", "")

if CLOUDINARY_CLOUD:
    cloudinary.config(
        cloud_name = CLOUDINARY_CLOUD,
        api_key    = CLOUDINARY_KEY,
        api_secret = CLOUDINARY_SEC,
        secure     = True
    )
    USE_CLOUDINARY = True
    print(f"✅ Cloudinary configurado: {CLOUDINARY_CLOUD}")
else:
    USE_CLOUDINARY = False
    print("⚠️  Cloudinary NO configurado — usando almacenamiento local")

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def _enviar_push(subscription_info, payload):
    """Envía una notificación push a una suscripción."""
    try:
        pk = VAPID_PRIVATE_KEY
        if not pk.startswith("-----"):
            pk_bytes = base64.urlsafe_b64decode(pk + "==")
            pk = pk_bytes.decode("utf-8")

        webpush(
            subscription_info=subscription_info,
            data=_json.dumps(payload),
            vapid_private_key=pk,
            vapid_claims={"sub": "mailto:" + VAPID_EMAIL}
        )
        return True
    except WebPushException as e:
        if "410" in str(e) or "404" in str(e):
            execute("DELETE FROM push_subscriptions WHERE endpoint=?", (subscription_info.get("endpoint",""),))
        else:
            print(f"⚠️ Error de push notification: {e}")
        return False
    except Exception as e:
        print(f"⚠️ Error inesperado al enviar push: {type(e).__name__}: {e}")
        return False


def notificar_cadetes_push(pedido_id, nombre_local, total, direccion_entrega):
    """Envía push a todos los cadetes disponibles y aprobados."""
    subs = query("""
        SELECT ps.endpoint, ps.p256dh, ps.auth
        FROM push_subscriptions ps
        JOIN usuarios u ON u.id = ps.usuario_id
        JOIN cadetes c ON c.usuario_id = u.id
        WHERE c.estado = 'aprobado' AND c.disponible = 1
    """)
    payload = {
        "titulo": "🛵 Nuevo pedido disponible",
        "cuerpo": f"{nombre_local} · ${int(total)} · {direccion_entrega or 'Retiro en local'}",
        "url":    "/mi-panel-cadete",
    }
    for sub in subs:
        _enviar_push({
            "endpoint": sub["endpoint"],
            "keys": {"p256dh": sub["p256dh"], "auth": sub["auth"]}
        }, payload)


def notificar_cliente_push(pedido_id, estado, nombre_local):
    """Avisa al cliente (si está logueado y tiene push activado) que su
    pedido cambió de estado. Los pedidos de clientes anónimos (sin cuenta)
    no tienen a quién avisarle por acá, siguen dependiendo de WhatsApp."""
    textos = {
        "confirmado": ("✔️ Pedido confirmado", f"{nombre_local} confirmó tu pedido y lo está preparando."),
        "en_camino":  ("🛵 Tu pedido está en camino", f"{nombre_local} · el cadete ya salió a entregarlo."),
        "entregado":  ("✅ Pedido entregado", f"¡Buen provecho! Gracias por pedir en {nombre_local}."),
        "cancelado":  ("✕ Pedido cancelado", f"Tu pedido en {nombre_local} fue cancelado."),
    }
    if estado not in textos:
        return
    pedido = query("SELECT cliente_id FROM pedidos WHERE id = ?", (pedido_id,), one=True)
    if not pedido or not pedido["cliente_id"]:
        return
    subs = query("""
        SELECT endpoint, p256dh, auth FROM push_subscriptions WHERE usuario_id = ?
    """, (pedido["cliente_id"],))
    titulo, cuerpo = textos[estado]
    payload = {"titulo": titulo, "cuerpo": cuerpo, "url": f"/pedido/{pedido_id}"}
    for sub in subs:
        _enviar_push({
            "endpoint": sub["endpoint"],
            "keys": {"p256dh": sub["p256dh"], "auth": sub["auth"]}
        }, payload)


def notificar_cadete_pedido_cancelado(pedido_id, cadete_id, nombre_local):
    """Avisa al cadete que ya tenía asignado un pedido (lo estaba yendo a
    entregar) que el local lo canceló — para que no siga viaje al pedo."""
    if not cadete_id:
        return
    cadete = query("SELECT usuario_id FROM cadetes WHERE id = ?", (cadete_id,), one=True)
    if not cadete:
        return
    subs = query("""
        SELECT endpoint, p256dh, auth FROM push_subscriptions WHERE usuario_id = ?
    """, (cadete["usuario_id"],))
    payload = {
        "titulo": "✕ Pedido cancelado",
        "cuerpo": f"{nombre_local} canceló el pedido #{pedido_id}. No hace falta que sigas viaje.",
        "url":    "/mi-panel-cadete",
    }
    for sub in subs:
        _enviar_push({
            "endpoint": sub["endpoint"],
            "keys": {"p256dh": sub["p256dh"], "auth": sub["auth"]}
        }, payload)


# ── GEOCODING (para el mapa de seguimiento) ───────────────────────────────────
import urllib.request, urllib.parse

def geocodificar_direccion(direccion, ciudad):
    """Convierte una dirección de texto en (lat, lng) usando Nominatim
    (OpenStreetMap), que es gratuito y no requiere API key. Se usa una sola
    vez por local/pedido y el resultado se cachea en la base — por eso el
    volumen de llamadas es bajo y no hace falta contratar un servicio pago.
    Si falla (dirección rara, sin internet, timeout) devuelve None y
    simplemente no se muestra ese marcador en el mapa, sin romper nada."""
    if not direccion:
        return None
    try:
        q = f"{direccion}, {ciudad}, Argentina"
        url = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode({
            "q": q, "format": "json", "limit": 1
        })
        req = urllib.request.Request(url, headers={
            "User-Agent": "PediacaApp/1.0 (contacto@pediaca.ar)"
        })
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = _json.loads(resp.read().decode())
        if data:
            return float(data[0]["lat"]), float(data[0]["lon"])
    except Exception as e:
        print(f"⚠️  Geocoding falló para '{direccion}': {e}")
    return None


def _sql(sql):
    return sql.replace("?", "%s") if USE_POSTGRES else sql

# ── BASE DE DATOS ─────────────────────────────────────────────────────────────
def get_db():
    if "db" not in g:
        if USE_POSTGRES:
            g.db = psycopg2.connect(DATABASE_URL)
        else:
            import sqlite3 as _sq
            g.db = _sq.connect(DB_PATH)
            g.db.row_factory = _sq.Row
            g.db.execute("PRAGMA foreign_keys = ON")
    return g.db

@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db:
        db.close()

def query(sql, args=(), one=False):
    sql = _sql(sql)
    if USE_POSTGRES:
        cur = get_db().cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql, args)
        rv = [dict(r) for r in cur.fetchall()]
    else:
        cur = get_db().execute(sql, args)
        cols = [d[0] for d in cur.description] if cur.description else []
        rv = [dict(zip(cols, row)) for row in cur.fetchall()]
    return (rv[0] if rv else None) if one else rv

def execute(sql, args=()):
    sql = _sql(sql)
    db  = get_db()
    if USE_POSTGRES:
        cur = db.cursor()
        sql_upper = sql.strip().upper()
        if sql_upper.startswith("INSERT") and "RETURNING" not in sql_upper:
            sql = sql.rstrip().rstrip(";") + " RETURNING id"
            cur.execute(sql, args)
            row = cur.fetchone()
            last_id = row[0] if row else None
        else:
            cur.execute(sql, args)
            last_id = None
        db.commit()
        return last_id
    else:
        import sqlite3 as _sq
        cur = db.execute(sql, args)
        db.commit()
        return cur.lastrowid


def _init_configuraciones():
    """Crea la tabla configuraciones (personalización del hero) si no existe
    y carga sus valores por defecto. Se ejecuta una vez al arrancar la app,
    dentro de un app_context real para que get_db()/execute() funcionen."""
    id_col = "id SERIAL PRIMARY KEY" if USE_POSTGRES else "id INTEGER PRIMARY KEY AUTOINCREMENT"
    try:
        with app.app_context():
            execute(f"""
                CREATE TABLE IF NOT EXISTS configuraciones (
                    {id_col},
                    clave TEXT UNIQUE,
                    valor TEXT,
                    tipo TEXT DEFAULT 'text',
                    actualizado TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            defaults = [
                ('hero_activo', 'false', 'text'),
                ('hero_imagen_url', '', 'text'),
                # Carrusel de publicidad a marcas (auspiciantes). Arranca
                # DESACTIVADO por defecto — Cristian lo prende desde el panel
                # de admin cuando quiera mostrarlo/usarlo en una demo.
                ('anuncios_activo', 'false', 'text'),
                # Banners laterales (a los costados de la página, solo
                # pantallas anchas). Toggle independiente del carrusel de
                # arriba, también arranca DESACTIVADO.
                ('anuncios_lateral_activo', 'false', 'text'),
            ]
            for clave, valor, tipo in defaults:
                execute("""
                    INSERT INTO configuraciones (clave, valor, tipo)
                    VALUES (?, ?, ?)
                    ON CONFLICT (clave) DO NOTHING
                """, (clave, valor, tipo))
            print("✅ Tabla configuraciones verificada/creada")
    except Exception as e:
        print(f"⚠️ Error al crear tabla configuraciones: {e}")


_init_configuraciones()


def _init_ciudades():
    """Crea la tabla ciudades (para poder activar/agregar ciudades desde el
    panel de admin sin tocar código ni reiniciar el server) y la siembra con
    Rosario activa más las principales ciudades del país como 'Próximamente'."""
    id_col = "id SERIAL PRIMARY KEY" if USE_POSTGRES else "id INTEGER PRIMARY KEY AUTOINCREMENT"
    try:
        with app.app_context():
            execute(f"""
                CREATE TABLE IF NOT EXISTS ciudades (
                    {id_col},
                    nombre TEXT UNIQUE NOT NULL,
                    activa INTEGER NOT NULL DEFAULT 0,
                    orden  INTEGER NOT NULL DEFAULT 0
                )
            """)
            defaults = [
                ("Rosario",       1, 1),
                ("Buenos Aires",  0, 2),
                ("Córdoba",       0, 3),
                ("Mendoza",       0, 4),
                ("La Plata",      0, 5),
                ("Mar del Plata", 0, 6),
                ("Tucumán",       0, 7),
                ("Salta",         0, 8),
                ("Santa Fe",      0, 9),
            ]
            for nombre, activa, orden in defaults:
                execute("""
                    INSERT INTO ciudades (nombre, activa, orden)
                    VALUES (?, ?, ?)
                    ON CONFLICT (nombre) DO NOTHING
                """, (nombre, activa, orden))
            print("✅ Tabla ciudades verificada/creada")
    except Exception as e:
        print(f"⚠️ Error al crear tabla ciudades: {e}")


_init_ciudades()


def _init_favoritos():
    """Crea la tabla favoritos si no existe (para instalaciones que ya
    tenían la base creada antes de que se agregara esta función)."""
    id_col = "id SERIAL PRIMARY KEY" if USE_POSTGRES else "id INTEGER PRIMARY KEY AUTOINCREMENT"
    ts_now = "NOW()" if USE_POSTGRES else "CURRENT_TIMESTAMP"
    try:
        with app.app_context():
            execute(f"""
                CREATE TABLE IF NOT EXISTS favoritos (
                    {id_col},
                    usuario_id     INTEGER NOT NULL REFERENCES usuarios(id),
                    restaurante_id INTEGER NOT NULL REFERENCES restaurantes(id),
                    fecha          TIMESTAMP NOT NULL DEFAULT {ts_now},
                    UNIQUE(usuario_id, restaurante_id)
                )
            """)
            print("✅ Tabla favoritos verificada/creada")
    except Exception as e:
        print(f"⚠️ Error al crear tabla favoritos: {e}")


_init_favoritos()


def _init_migrar_columnas():
    """Agrega columnas nuevas a tablas que ya existen, sin depender de que
    alguien se acuerde de correr migrate_db.py a mano antes de levantar la
    app (eso ya causó un KeyError en producción/local una vez)."""
    columnas = [
        ("restaurantes", "abierto",              "INTEGER NOT NULL DEFAULT 1"),
        ("restaurantes", "ciudad",               "TEXT NOT NULL DEFAULT 'Rosario'"),
        ("promociones",  "producto_id",          "INTEGER REFERENCES productos(id)"),
        ("promociones",  "tipo_descuento",       "TEXT DEFAULT 'porcentaje'"),
        ("promociones",  "descuento_monto",      "REAL DEFAULT 0"),
        ("promociones",  "precio_con_descuento", "REAL"),
        ("cadetes",      "ciudad",               "TEXT NOT NULL DEFAULT 'Rosario'"),
        ("valoraciones", "vista",                "INTEGER NOT NULL DEFAULT 0"),
        ("pedidos",      "costo_envio",          "REAL NOT NULL DEFAULT 0"),
        ("pedidos",      "codigo_entrega",       "TEXT"),
    ]
    try:
        with app.app_context():
            for tabla, columna, tipo in columnas:
                try:
                    if USE_POSTGRES:
                        execute(f"ALTER TABLE {tabla} ADD COLUMN IF NOT EXISTS {columna} {tipo}")
                    else:
                        existentes = {row["name"] for row in query(f"PRAGMA table_info({tabla})")}
                        if columna in existentes:
                            continue
                        execute(f"ALTER TABLE {tabla} ADD COLUMN {columna} {tipo}")
                except Exception as e:
                    print(f"⚠️ Migración columna {tabla}.{columna}: {e}")
            print("✅ Columnas verificadas/migradas")
    except Exception as e:
        print(f"⚠️ Error en migración de columnas: {e}")


_init_migrar_columnas()


@app.context_processor
def inject_now_year():
    from datetime import datetime
    return dict(now_year=datetime.now().year)


# ── HELPERS ───────────────────────────────────────────────────────────────────
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXT

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            flash("Necesitás iniciar sesión.", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def rol_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if session.get("rol") not in roles:
                flash("No tenés permiso para acceder a esa página.", "danger")
                return redirect(url_for("home"))
            return f(*args, **kwargs)
        return decorated
    return decorator

def formato_pesos(valor):
    return f"${valor:,.0f}".replace(",", ".")

def formato_fecha(valor, formato="%Y-%m-%d %H:%M"):
    if not valor:
        return ""
    if hasattr(valor, 'strftime'):
        return valor.strftime(formato)
    return str(valor)[:len(formato.replace("%Y","0000").replace("%m","00").replace("%d","00").replace("%H","00").replace("%M","00"))]

app.jinja_env.filters["pesos"]  = formato_pesos
app.jinja_env.filters["fecha"]  = formato_fecha
app.jinja_env.filters["fecha_corta"] = lambda v: formato_fecha(v, "%Y-%m-%d")


# ── CIUDADES ──────────────────────────────────────────────────────────────────
# Se administran desde /admin/configuracion (tabla `ciudades`), así se pueden
# activar ciudades nuevas o agregar otras sin tocar código ni reiniciar el server.
CIUDAD_DEFAULT = "Rosario"


def get_ciudades():
    return query("SELECT id, nombre, activa FROM ciudades ORDER BY orden, nombre")


def get_ciudades_validas():
    return {c["nombre"] for c in get_ciudades()}


def get_ciudad_actual():
    ciudad = request.cookies.get("ciudad", CIUDAD_DEFAULT)
    return ciudad if ciudad in get_ciudades_validas() else CIUDAD_DEFAULT


@app.context_processor
def inject_ciudad():
    return dict(ciudades_disponibles=get_ciudades(), ciudad_actual=get_ciudad_actual())


@app.route("/ciudad/<ciudad>")
def cambiar_ciudad(ciudad):
    resp = redirect(request.referrer or url_for("home"))
    if ciudad in get_ciudades_validas():
        resp.set_cookie("ciudad", ciudad, max_age=60 * 60 * 24 * 365)
    return resp


def _precio_con_promo(precio, promo):
    """Calcula el precio final de un producto con una promo aplicada."""
    if not promo:
        return None
    if promo.get("precio_con_descuento"):
        final = promo["precio_con_descuento"]
    elif promo.get("tipo_descuento") == "porcentaje" and promo.get("descuento_pct"):
        final = precio * (1 - promo["descuento_pct"] / 100)
    elif promo.get("tipo_descuento") == "monto" and promo.get("descuento_monto"):
        final = precio - promo["descuento_monto"]
    else:
        return None
    final = max(0, round(final))
    return final if final < precio else None


# ── RUTAS PÚBLICAS ────────────────────────────────────────────────────────────

@app.route("/")
def home():
    ciudad = get_ciudad_actual()

    restaurantes = query("""
        SELECT r.*, u.telefono,
               COALESCE(AVG(v.estrellas), 0) AS rating,
               COUNT(DISTINCT p.id) AS total_pedidos
        FROM restaurantes r
        JOIN usuarios u ON u.id = r.usuario_id
        LEFT JOIN valoraciones v ON v.restaurante_id = r.id
        LEFT JOIN pedidos p ON p.restaurante_id = r.id AND p.estado = 'entregado'
        WHERE r.estado = 'aprobado' AND r.ciudad = ?
        GROUP BY r.id, u.telefono
        ORDER BY r.abierto DESC, r.nombre_local
    """, (ciudad,))
    anuncios_activo = get_config("anuncios_activo") == "true"
    auspiciantes = query("""
        SELECT * FROM auspiciantes
        WHERE activo = 1
          AND (fecha_inicio IS NULL OR fecha_inicio <= CURRENT_DATE)
          AND (fecha_fin   IS NULL OR fecha_fin   >= CURRENT_DATE)
    """) if anuncios_activo else []

    promociones_destacadas = query("""
        SELECT p.*, r.nombre_local, r.id as restaurante_id
        FROM promociones p
        JOIN restaurantes r ON r.id = p.restaurante_id
        WHERE p.activa = 1
          AND r.estado = 'aprobado'
          AND r.ciudad = ?
        ORDER BY p.fecha_creacion DESC
        LIMIT 10
    """, (ciudad,))

    favoritos_ids = set()
    if session.get("rol") == "cliente":
        favs = query("SELECT restaurante_id FROM favoritos WHERE usuario_id=?", (session["user_id"],))
        favoritos_ids = {f["restaurante_id"] for f in favs}

    return render_template("home.html", restaurantes=restaurantes, auspiciantes=auspiciantes,
                           anuncios_activo=anuncios_activo,
                           promociones_destacadas=promociones_destacadas,
                           favoritos_ids=favoritos_ids)


@app.route("/local/<int:restaurante_id>")
def ver_local(restaurante_id):
    restaurante = query("""
        SELECT r.*, u.nombre AS dueno_nombre
        FROM restaurantes r
        JOIN usuarios u ON u.id = r.usuario_id
        WHERE r.id = ? AND r.estado = 'aprobado'
    """, (restaurante_id,), one=True)

    if not restaurante:
        flash("Local no encontrado.", "danger")
        return redirect(url_for("home"))

    categorias = query("""
        SELECT * FROM categorias_menu
        WHERE restaurante_id = ?
        ORDER BY orden
    """, (restaurante_id,))

    productos = query("""
        SELECT * FROM productos
        WHERE restaurante_id = ? AND disponible = 1
        ORDER BY categoria_id, orden
    """, (restaurante_id,))

    valoraciones_pub = query("""
        SELECT v.estrellas, v.comentario, v.fecha
        FROM valoraciones v
        WHERE v.restaurante_id = ?
        ORDER BY v.fecha DESC LIMIT 10
    """, (restaurante_id,))

    rating_avg = 0
    if valoraciones_pub:
        rating_avg = sum(v["estrellas"] for v in valoraciones_pub) / len(valoraciones_pub)

    promociones = query("""
        SELECT * FROM promociones
        WHERE restaurante_id = ? AND activa = 1
          AND (fecha_inicio IS NULL OR fecha_inicio <= CURRENT_DATE)
          AND (fecha_fin   IS NULL OR fecha_fin   >= CURRENT_DATE)
        ORDER BY fecha_creacion DESC
    """, (restaurante_id,))

    # Promos ligadas a un producto puntual → se descuentan solas en el
    # precio de ese producto. Promos sin producto (generales del local) →
    # se aplican como descuento sobre el total del pedido en el carrito.
    promo_por_producto = {}
    promo_orden = None
    for promo in promociones:
        if promo.get("producto_id"):
            promo_por_producto[promo["producto_id"]] = promo
        elif promo_orden is None:
            promo_orden = promo

    for p in productos:
        promo = promo_por_producto.get(p["id"])
        p["precio_promo"] = _precio_con_promo(p["precio"], promo) if promo else None

    # Productos sin categoría asignada (o con categoria_id huérfano, apuntando
    # a una categoría borrada) quedaban invisibles en el menú público, porque
    # el template solo recorre las categorías existentes y nunca mostraba un
    # catch-all para estos casos. Los normalizamos a categoria_id=None acá y
    # les creamos una categoría sintética "Otros" para que siempre se vean.
    cat_ids = {c["id"] for c in categorias}
    for p in productos:
        if p["categoria_id"] not in cat_ids:
            p["categoria_id"] = None
    if any(p["categoria_id"] is None for p in productos):
        categorias = categorias + [{"id": None, "nombre": "Otros", "orden": 999999}]

    es_favorito = False
    if session.get("rol") == "cliente":
        fav = query("SELECT id FROM favoritos WHERE usuario_id=? AND restaurante_id=?",
                    (session["user_id"], restaurante_id), one=True)
        es_favorito = bool(fav)

    return render_template("ver_local.html",
                           promo_orden=promo_orden,
                           restaurante=restaurante,
                           categorias=categorias,
                           productos=productos,
                           valoraciones=valoraciones_pub,
                           rating_avg=rating_avg,
                           promociones=promociones,
                           es_favorito=es_favorito)

# ── REGISTRO ──────────────────────────────────────────────────────────────────



@app.route("/promociones")
def ver_promociones():
    """Muestra todas las promociones activas de todos los locales"""
    promociones = query("""
        SELECT p.*, r.nombre_local, r.id as restaurante_id, r.logo_url, r.categoria
        FROM promociones p
        JOIN restaurantes r ON r.id = p.restaurante_id
        WHERE p.activa = 1
          AND r.estado = 'aprobado'
          AND r.abierto = 1
        ORDER BY p.fecha_creacion DESC
    """)
    
    # Agrupar por tipo de descuento para mostrar
    for promo in promociones:
        if promo['tipo_descuento'] == 'porcentaje' and promo['descuento_pct']:
            promo['descuento_texto'] = f"{promo['descuento_pct']}% OFF"
        elif promo['descuento_monto']:
            promo['descuento_texto'] = f"${promo['descuento_monto']} OFF"
        else:
            promo['descuento_texto'] = "Oferta especial"
    
    return render_template("promociones_lista.html", promociones=promociones)

@app.route("/registro", methods=["GET", "POST"])
def registro():
    if request.method == "POST":
        nombre    = request.form.get("nombre", "").strip()
        apellido  = request.form.get("apellido", "").strip()
        email     = request.form.get("email", "").strip().lower()
        telefono  = request.form.get("telefono", "").strip()
        password  = request.form.get("password", "")
        password2 = request.form.get("password2", "")
        rol       = request.form.get("rol", "cliente")

        if not all([nombre, apellido, email, password]):
            flash("Completá todos los campos obligatorios.", "danger")
            return redirect(url_for("registro"))

        if password != password2:
            flash("Las contraseñas no coinciden.", "danger")
            return redirect(url_for("registro"))

        if len(password) < 6:
            flash("La contraseña debe tener al menos 6 caracteres.", "danger")
            return redirect(url_for("registro"))

        if rol not in ("cliente", "restaurante", "cadete"):
            flash("Rol inválido.", "danger")
            return redirect(url_for("registro"))

        existe = query("SELECT id FROM usuarios WHERE email = ?", (email,), one=True)
        if existe:
            flash("Ya existe una cuenta con ese email.", "warning")
            return redirect(url_for("registro"))

        password_hash = generate_password_hash(password)
        user_id = execute("""
            INSERT INTO usuarios (nombre, apellido, email, telefono, password_hash, rol)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (nombre, apellido, email, telefono, password_hash, rol))

        if rol == "cliente":
            execute("INSERT INTO clientes (usuario_id) VALUES (?)", (user_id,))
            flash("¡Cuenta creada! Ya podés iniciar sesión.", "success")

        elif rol == "restaurante":
            nombre_local = request.form.get("nombre_local", "").strip()
            whatsapp     = request.form.get("whatsapp", "").strip()
            categoria    = request.form.get("categoria", "").strip()
            direccion    = request.form.get("direccion", "").strip()
            ciudad       = request.form.get("ciudad", "").strip()
            if ciudad not in get_ciudades_validas():
                ciudad = CIUDAD_DEFAULT

            if not nombre_local or not whatsapp:
                flash("El nombre del local y el WhatsApp son obligatorios.", "danger")
                execute("DELETE FROM usuarios WHERE id = ?", (user_id,))
                return redirect(url_for("registro"))

            execute("""
                INSERT INTO restaurantes (usuario_id, nombre_local, whatsapp, categoria, direccion, ciudad)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (user_id, nombre_local, whatsapp, categoria, direccion, ciudad))
            flash("Registro enviado. Te avisamos cuando tu local esté aprobado.", "info")

        elif rol == "cadete":
            vehiculo    = request.form.get("vehiculo", "moto")
            zona        = request.form.get("zona", "").strip()
            ciudad_cad  = request.form.get("ciudad", "").strip()
            if ciudad_cad not in get_ciudades_validas():
                ciudad_cad = CIUDAD_DEFAULT
            execute("""
                INSERT INTO cadetes (usuario_id, vehiculo, zona, ciudad)
                VALUES (?, ?, ?, ?)
            """, (user_id, vehiculo, zona, ciudad_cad))
            flash("Registro enviado. Te avisamos cuando seas aprobado.", "info")

        return redirect(url_for("login"))

    return render_template("registro.html")


# ── LOGIN / LOGOUT ────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        email    = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        user = query("SELECT * FROM usuarios WHERE email = ?", (email,), one=True)

        if not user or not check_password_hash(user["password_hash"], password):
            flash("Email o contraseña incorrectos.", "danger")
            return redirect(url_for("login"))

        if not user["activo"]:
            flash("Tu cuenta está desactivada.", "warning")
            return redirect(url_for("login"))

        session["user_id"] = user["id"]
        session["nombre"]  = user["nombre"]
        session["rol"]     = user["rol"]

        flash(f"¡Bienvenido, {user['nombre']}!", "success")
        return redirect(url_for("dashboard"))

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("home"))


# ── DASHBOARD ─────────────────────────────────────────────────────────────────

@app.route("/dashboard")
@login_required
def dashboard():
    rol = session.get("rol")
    if rol == "admin":
        return redirect(url_for("admin_panel"))
    if rol == "restaurante":
        return redirect(url_for("restaurante_panel"))
    if rol == "cadete":
        return redirect(url_for("cadete_panel"))
    return redirect(url_for("cliente_panel"))


# ── PANEL RESTAURANTE ─────────────────────────────────────────────────────────

def get_restaurante_aprobado():
    return query("""
        SELECT * FROM restaurantes
        WHERE usuario_id = ? AND estado = 'aprobado'
    """, (session["user_id"],), one=True)

def get_restaurante_any():
    return query(
        "SELECT * FROM restaurantes WHERE usuario_id = ?",
        (session["user_id"],), one=True
    )


@app.route("/sw.js")
def service_worker():
    """Sirve el service worker en la raíz (no en /static/) para que su
    scope sea todo el sitio y la PWA sea instalable en cualquier página."""
    resp = app.send_static_file("sw.js")
    resp.headers["Cache-Control"] = "no-cache"
    return resp


@app.route("/api/vapid-public-key")
def vapid_public_key():
    return jsonify({"publicKey": VAPID_PUBLIC_KEY})


@app.route("/api/push/subscribe", methods=["POST"])
@login_required
def push_subscribe():
    data     = request.get_json()
    endpoint = data.get("endpoint")
    p256dh   = data.get("keys", {}).get("p256dh")
    auth     = data.get("keys", {}).get("auth")
    if not all([endpoint, p256dh, auth]):
        return jsonify({"error": "Datos incompletos"}), 400
    existe = query("SELECT id FROM push_subscriptions WHERE endpoint=?", (endpoint,), one=True)
    if existe:
        execute("UPDATE push_subscriptions SET p256dh=?, auth=?, usuario_id=? WHERE endpoint=?",
                (p256dh, auth, session["user_id"], endpoint))
    else:
        execute("""INSERT INTO push_subscriptions (usuario_id, endpoint, p256dh, auth)
                   VALUES (?,?,?,?)""", (session["user_id"], endpoint, p256dh, auth))
    return jsonify({"ok": True})


@app.route("/api/push/unsubscribe", methods=["POST"])
@login_required
def push_unsubscribe():
    data = request.get_json()
    execute("DELETE FROM push_subscriptions WHERE endpoint=? AND usuario_id=?",
            (data.get("endpoint", ""), session["user_id"]))
    return jsonify({"ok": True})


def guardar_imagen(archivo, subcarpeta=""):
    if not archivo or archivo.filename == "":
        return None
    if not allowed_file(archivo.filename):
        return None

    if USE_CLOUDINARY:
        try:
            resultado = cloudinary.uploader.upload(
                archivo,
                folder=f"pediaca/{subcarpeta}",
                resource_type="image",
                quality="auto",
                fetch_format="auto",
            )
            return resultado["secure_url"]
        except Exception as e:
            print(f"❌ Error Cloudinary: {type(e).__name__}: {e}")

    import uuid, time
    filename = secure_filename(archivo.filename)
    ext      = filename.rsplit(".", 1)[1].lower()
    filename = f"{int(time.time())}_{uuid.uuid4().hex[:6]}.{ext}"
    destino  = os.path.join(app.config["UPLOAD_FOLDER"], subcarpeta)
    os.makedirs(destino, exist_ok=True)
    archivo.save(os.path.join(destino, filename))
    return os.path.join("uploads", subcarpeta, filename).replace("\\", "/")


def url_imagen(ruta):
    if not ruta:
        return None
    if ruta.startswith("http"):
        return ruta
    return url_for("static", filename=ruta)


app.jinja_env.globals["url_imagen"] = url_imagen


@app.route("/mi-local")
@login_required
@rol_required("restaurante")
def restaurante_panel():
    restaurante_raw = get_restaurante_any()
    if not restaurante_raw:
        flash("No encontramos tu local.", "danger")
        return redirect(url_for("home"))

    if restaurante_raw["estado"] != "aprobado":
        return render_template("restaurante_espera.html",
                               restaurante=restaurante_raw)

    restaurante = restaurante_raw

    categorias = query("""
        SELECT c.*, COUNT(p.id) AS total_productos
        FROM categorias_menu c
        LEFT JOIN productos p ON p.categoria_id = c.id
        WHERE c.restaurante_id = ?
        GROUP BY c.id
        ORDER BY c.orden
    """, (restaurante["id"],))

    productos = query("""
        SELECT p.*, c.nombre AS categoria_nombre
        FROM productos p
        LEFT JOIN categorias_menu c ON c.id = p.categoria_id
        WHERE p.restaurante_id = ?
        ORDER BY p.categoria_id, p.orden
    """, (restaurante["id"],))

    sabores_map = {}
    if productos:
        ids = ",".join(str(p["id"]) for p in productos)
        sabores = query(f"SELECT * FROM sabores_producto WHERE producto_id IN ({ids}) AND disponible=1 ORDER BY producto_id, orden")
        for s in sabores:
            sabores_map.setdefault(s["producto_id"], []).append(s)

    valoraciones = query("""
        SELECT v.*, p.nombre_cliente_anonimo
        FROM valoraciones v
        LEFT JOIN pedidos p ON p.id = v.pedido_id
        WHERE v.restaurante_id = ?
        ORDER BY v.fecha DESC
    """, (restaurante["id"],))
    valoraciones_no_vistas = sum(1 for v in valoraciones if not v.get("vista"))

    promociones = query("""
        SELECT * FROM promociones
        WHERE restaurante_id = ?
        ORDER BY fecha_creacion DESC
    """, (restaurante["id"],))

    return render_template("restaurante_panel.html",
                           restaurante=restaurante,
                           categorias=categorias,
                           productos=productos,
                           sabores_map=sabores_map,
                           valoraciones=valoraciones,
                           valoraciones_no_vistas=valoraciones_no_vistas,
                           promociones=promociones)


@app.route("/mi-local/valoraciones/marcar-vistas", methods=["POST"])
@login_required
@rol_required("restaurante")
def marcar_valoraciones_vistas():
    restaurante = get_restaurante_any()
    if restaurante:
        execute("UPDATE valoraciones SET vista=1 WHERE restaurante_id=?", (restaurante["id"],))
    return jsonify({"ok": True})


@app.route("/mi-local/editar", methods=["POST"])
@login_required
@rol_required("restaurante")
def restaurante_editar():
    restaurante = query(
        "SELECT * FROM restaurantes WHERE usuario_id = ?",
        (session["user_id"],), one=True
    )
    if not restaurante:
        return redirect(url_for("home"))

    nombre_local    = request.form.get("nombre_local", "").strip()
    descripcion     = request.form.get("descripcion", "").strip()
    categoria       = request.form.get("categoria", "").strip()
    direccion       = request.form.get("direccion", "").strip()
    whatsapp        = request.form.get("whatsapp", "").strip()
    horario         = request.form.get("horario", "").strip()
    hace_envio      = 1 if request.form.get("hace_envio") else 0
    costo_envio     = float(request.form.get("costo_envio", 0) or 0)
    tiempo_estimado = request.form.get("tiempo_estimado", "").strip() or None
    ciudad          = request.form.get("ciudad", "").strip()
    if ciudad not in get_ciudades_validas():
        ciudad = restaurante["ciudad"] or CIUDAD_DEFAULT

    execute("""
        UPDATE restaurantes SET
            nombre_local=?, descripcion=?, categoria=?, direccion=?,
            whatsapp=?, horario=?, hace_envio=?, costo_envio=?, tiempo_estimado=?, ciudad=?
        WHERE id=?
    """, (nombre_local, descripcion, categoria, direccion,
          whatsapp, horario, hace_envio, costo_envio,
          tiempo_estimado, ciudad, restaurante["id"]))

    flash("Datos del local actualizados.", "success")
    return redirect(url_for("restaurante_panel"))


@app.route("/mi-local/categoria/nueva", methods=["POST"])
@login_required
@rol_required("restaurante")
def categoria_nueva():
    restaurante = query(
        "SELECT id FROM restaurantes WHERE usuario_id = ?",
        (session["user_id"],), one=True
    )
    nombre = request.form.get("nombre", "").strip()
    if nombre and restaurante:
        execute("""
            INSERT INTO categorias_menu (restaurante_id, nombre, orden)
            VALUES (?, ?, (SELECT COALESCE(MAX(orden),0)+1 FROM categorias_menu WHERE restaurante_id=?))
        """, (restaurante["id"], nombre, restaurante["id"]))
        flash(f"Categoría '{nombre}' creada.", "success")
    return redirect(url_for("restaurante_panel"))


@app.route("/mi-local/producto/nuevo", methods=["POST"])
@login_required
@rol_required("restaurante")
def producto_nuevo():
    restaurante = query(
        "SELECT id FROM restaurantes WHERE usuario_id = ?",
        (session["user_id"],), one=True
    )
    if not restaurante:
        return redirect(url_for("home"))

    nombre       = request.form.get("nombre", "").strip()
    descripcion  = request.form.get("descripcion", "").strip()
    precio       = float(request.form.get("precio", 0) or 0)
    categoria_id = request.form.get("categoria_id") or None
    disponible   = 1 if request.form.get("disponible") else 0
    sabores_por_kilo = None
    if request.form.get("vendido_por_peso"):
        sabores_por_kilo = int(request.form.get("sabores_por_kilo", 0) or 0) or None

    foto_url = None
    archivo  = request.files.get("foto")
    if archivo and archivo.filename:
        foto_url = guardar_imagen(archivo, "productos")

    prod_id = execute("""
        INSERT INTO productos (restaurante_id, categoria_id, nombre, descripcion, precio, disponible, foto_url, sabores_por_kilo)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (restaurante["id"], categoria_id, nombre, descripcion, precio, disponible, foto_url, sabores_por_kilo))

    flash(f"Producto '{nombre}' agregado.", "success")
    return redirect(url_for("restaurante_panel"))


# ── CARGA MASIVA ───────────────────────────────────────────────────────────────

@app.route("/mi-local/carga-masiva", methods=["POST"])
@login_required
@rol_required("restaurante")
def carga_masiva():
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        return jsonify({"error": "No autorizado"}), 403

    productos_data = []

    if request.is_json:
        rows = request.get_json().get("productos", [])
        for r in rows:
            nombre = str(r.get("nombre", "")).strip()
            if not nombre:
                continue
            productos_data.append({
                "nombre":      nombre,
                "descripcion": str(r.get("descripcion", "")).strip(),
                "precio":      float(r.get("precio", 0) or 0),
                "categoria":   str(r.get("categoria", "")).strip(),
            })

    else:
        archivo = request.files.get("archivo")
        if not archivo:
            return jsonify({"error": "No se recibió archivo"}), 400

        ext = archivo.filename.rsplit(".", 1)[-1].lower()

        if ext == "csv":
            import csv, io as _io
            texto = archivo.read().decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(_io.StringIO(texto))
            for row in reader:
                nombre = str(row.get("nombre", row.get("Nombre", ""))).strip()
                if not nombre:
                    continue
                productos_data.append({
                    "nombre":      nombre,
                    "descripcion": str(row.get("descripcion", row.get("Descripcion", ""))).strip(),
                    "precio":      float(row.get("precio", row.get("Precio", 0)) or 0),
                    "categoria":   str(row.get("categoria", row.get("Categoria", ""))).strip(),
                })

        elif ext in ("xlsx", "xls"):
            import openpyxl
            wb   = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return jsonify({"error": "Archivo vacío"}), 400
            headers = [str(h or "").lower().strip() for h in rows[0]]
            def _col(row, *names):
                for n in names:
                    if n in headers:
                        v = row[headers.index(n)]
                        return v if v is not None else ""
                return ""
            for row in rows[1:]:
                nombre = str(_col(row, "nombre")).strip()
                if not nombre:
                    continue
                productos_data.append({
                    "nombre":      nombre,
                    "descripcion": str(_col(row, "descripcion")).strip(),
                    "precio":      float(_col(row, "precio") or 0),
                    "categoria":   str(_col(row, "categoria")).strip(),
                })
        else:
            return jsonify({"error": "Formato no soportado. Usá .csv o .xlsx"}), 400

    if not productos_data:
        return jsonify({"error": "No se encontraron productos válidos"}), 400

    cats = query("SELECT id, nombre FROM categorias_menu WHERE restaurante_id=?",
                 (restaurante["id"],))
    cat_map = {c["nombre"].lower().strip(): c["id"] for c in cats}

    insertados = 0
    categorias_creadas = 0
    for p in productos_data:
        cat_id = None
        cat_nombre = p["categoria"].strip()
        if cat_nombre:
            cat_id = cat_map.get(cat_nombre.lower())
            if cat_id is None:
                # La categoría no existe todavía para este local (caso típico:
                # carga masiva de un menú nuevo) — la creamos en vez de dejar
                # el producto sin categoría silenciosamente.
                cat_id = execute("""
                    INSERT INTO categorias_menu (restaurante_id, nombre)
                    VALUES (?, ?)
                """, (restaurante["id"], cat_nombre))
                cat_map[cat_nombre.lower()] = cat_id
                categorias_creadas += 1
        execute("""
            INSERT INTO productos (restaurante_id, categoria_id, nombre, descripcion, precio, disponible)
            VALUES (?, ?, ?, ?, ?, 1)
        """, (restaurante["id"], cat_id, p["nombre"], p["descripcion"], p["precio"]))
        insertados += 1

    return jsonify({"ok": True, "insertados": insertados, "categorias_creadas": categorias_creadas})


@app.route("/mi-local/plantilla-excel")
@login_required
@rol_required("restaurante")
def descargar_plantilla():
    import openpyxl
    from flask import send_file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Menú"
    ws.append(["nombre", "descripcion", "precio", "categoria"])
    ejemplos = [
        ("Pizza Muzzarella", "Salsa, muzzarella y orégano", 4500, "Pizzas"),
        ("Pizza Napolitana", "Salsa, muzzarella y tomate", 5000, "Pizzas"),
        ("Empanada de Carne", "Carne cortada a cuchillo", 800, "Empanadas"),
        ("Coca-Cola 500ml", "", 1200, "Bebidas"),
    ]
    for e in ejemplos:
        ws.append(e)
    for col, ancho in [("A", 30), ("B", 40), ("C", 12), ("D", 20)]:
        ws.column_dimensions[col].width = ancho

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="plantilla_menu_pediaca.xlsx")


# ── SABORES ────────────────────────────────────────────────────────────────────

@app.route("/mi-local/producto/<int:prod_id>/sabores", methods=["POST"])
@login_required
@rol_required("restaurante")
def sabor_nuevo(prod_id):
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        return redirect(url_for("restaurante_panel"))
    prod = query("SELECT id FROM productos WHERE id=? AND restaurante_id=?",
                 (prod_id, restaurante["id"]), one=True)
    if not prod:
        return redirect(url_for("restaurante_panel"))

    nombre = request.form.get("nombre", "").strip()
    if nombre:
        execute("""
            INSERT INTO sabores_producto (producto_id, nombre, orden)
            VALUES (?, ?, (SELECT COALESCE(MAX(orden),0)+1 FROM sabores_producto WHERE producto_id=?))
        """, (prod_id, nombre, prod_id))
        flash(f"Gusto '{nombre}' agregado.", "success")
    return redirect(url_for("restaurante_panel") + "#prod-" + str(prod_id))


@app.route("/mi-local/sabor/<int:sabor_id>/eliminar", methods=["POST"])
@login_required
@rol_required("restaurante")
def sabor_eliminar(sabor_id):
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        return redirect(url_for("restaurante_panel"))
    execute("""
        DELETE FROM sabores_producto
        WHERE id=? AND producto_id IN (
            SELECT id FROM productos WHERE restaurante_id=?
        )
    """, (sabor_id, restaurante["id"]))
    return redirect(url_for("restaurante_panel"))


@app.route("/api/producto/<int:prod_id>/sabores")
def get_sabores(prod_id):
    sabores = query("""
        SELECT id, nombre FROM sabores_producto
        WHERE producto_id=? AND disponible=1
        ORDER BY orden
    """, (prod_id,))
    return jsonify({"sabores": [dict(s) for s in sabores]})


@app.route("/mi-local/producto/<int:prod_id>/toggle", methods=["POST"])
@login_required
@rol_required("restaurante")
def producto_toggle(prod_id):
    restaurante = query(
        "SELECT id FROM restaurantes WHERE usuario_id = ?",
        (session["user_id"],), one=True
    )
    if not restaurante:
        return redirect(url_for("restaurante_panel"))
    execute("""
        UPDATE productos SET disponible = 1 - disponible
        WHERE id = ? AND restaurante_id = ?
    """, (prod_id, restaurante["id"]))
    return redirect(url_for("restaurante_panel"))


@app.route("/mi-local/producto/<int:prod_id>/eliminar", methods=["POST"])
@login_required
@rol_required("restaurante")
def producto_eliminar(prod_id):
    restaurante = query(
        "SELECT id FROM restaurantes WHERE usuario_id = ?",
        (session["user_id"],), one=True
    )
    if not restaurante:
        return redirect(url_for("restaurante_panel"))
    execute("DELETE FROM productos WHERE id = ? AND restaurante_id = ?",
            (prod_id, restaurante["id"]))
    flash("Producto eliminado.", "success")
    return redirect(url_for("restaurante_panel"))


# ── PRODUCTO EDITAR ───────────────────────────────────────────────────────────

@app.route("/mi-local/producto/<int:prod_id>/editar", methods=["POST"])
@login_required
@rol_required("restaurante")
def producto_editar(prod_id):
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        return redirect(url_for("restaurante_panel"))

    nombre       = request.form.get("nombre", "").strip()
    descripcion  = request.form.get("descripcion", "").strip()
    precio       = float(request.form.get("precio", 0) or 0)
    categoria_id = request.form.get("categoria_id") or None
    disponible   = 1 if request.form.get("disponible") else 0
    sabores_por_kilo = None
    if request.form.get("vendido_por_peso"):
        sabores_por_kilo = int(request.form.get("sabores_por_kilo", 0) or 0) or None

    execute("""
        UPDATE productos SET
            nombre=?, descripcion=?, precio=?, categoria_id=?, disponible=?, sabores_por_kilo=?
        WHERE id=? AND restaurante_id=?
    """, (nombre, descripcion, precio, categoria_id, disponible, sabores_por_kilo,
          prod_id, restaurante["id"]))
    flash("Producto actualizado.", "success")
    return redirect(url_for("restaurante_panel"))


# ── PANEL CADETE ──────────────────────────────────────────────────────────────

@app.route("/mi-panel-cadete")
@login_required
@rol_required("cadete")
def cadete_panel():
    cadete = query(
        "SELECT * FROM cadetes WHERE usuario_id = ?",
        (session["user_id"],), one=True
    )
    pedidos_disponibles = []
    entregas = []
    entregas_stats = {"hoy": 0, "semana": 0, "total": 0}
    if cadete and cadete["estado"] == "aprobado":
        pedidos_disponibles = query("""
            SELECT p.*, r.nombre_local, r.whatsapp
            FROM pedidos p
            JOIN restaurantes r ON r.id = p.restaurante_id
            WHERE p.tipo_entrega = 'delivery'
              AND p.estado = 'confirmado'
              AND p.cadete_id IS NULL
              AND r.ciudad = ?
            ORDER BY p.fecha_pedido DESC
        """, (cadete.get("ciudad") or CIUDAD_DEFAULT,))

        entregas = query("""
            SELECT p.id, p.total, p.costo_envio, p.fecha_pedido, p.direccion_entrega, r.nombre_local
            FROM pedidos p
            JOIN restaurantes r ON r.id = p.restaurante_id
            WHERE p.cadete_id = ? AND p.estado = 'entregado'
            ORDER BY p.fecha_pedido DESC LIMIT 50
        """, (cadete["id"],))

        from datetime import datetime, timedelta
        hace_7_dias = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')
        entregas_stats["total"] = query(
            "SELECT COUNT(*) as n FROM pedidos WHERE cadete_id=? AND estado='entregado'",
            (cadete["id"],), one=True)["n"]
        entregas_stats["semana"] = query(
            "SELECT COUNT(*) as n FROM pedidos WHERE cadete_id=? AND estado='entregado' AND fecha_pedido >= ?",
            (cadete["id"], hace_7_dias), one=True)["n"]
        entregas_stats["hoy"] = query(
            "SELECT COUNT(*) as n FROM pedidos WHERE cadete_id=? AND estado='entregado' AND date(fecha_pedido)=CURRENT_DATE",
            (cadete["id"],), one=True)["n"]

    entrega_en_curso = None
    if cadete and cadete["estado"] == "aprobado":
        # Si el cadete tiene una entrega "en camino" asignada, el navegador
        # empieza a mandar su ubicación GPS (ver JS) para que el comprador y
        # el local la vean en el mapa de seguimiento.
        entrega_en_curso = query("""
            SELECT p.id, p.total, p.costo_envio, p.direccion_entrega,
                   r.nombre_local, r.direccion AS local_direccion
            FROM pedidos p JOIN restaurantes r ON r.id = p.restaurante_id
            WHERE p.cadete_id = ? AND p.estado = 'en_camino'
            ORDER BY p.fecha_pedido DESC LIMIT 1
        """, (cadete["id"],), one=True)

    return render_template("cadete_panel.html",
                           cadete=cadete,
                           pedidos=pedidos_disponibles,
                           entregas=entregas,
                           entregas_stats=entregas_stats,
                           entrega_en_curso=entrega_en_curso)


@app.route("/mi-panel-cadete/disponibilidad", methods=["POST"])
@login_required
@rol_required("cadete")
def cadete_toggle_disponibilidad():
    cadete = query("SELECT * FROM cadetes WHERE usuario_id = ?",
                   (session["user_id"],), one=True)
    if cadete:
        execute("UPDATE cadetes SET disponible = 1 - disponible WHERE usuario_id = ?",
                (session["user_id"],))
    return redirect(url_for("cadete_panel"))


@app.route("/mi-panel-cadete/editar", methods=["POST"])
@login_required
@rol_required("cadete")
def cadete_editar_perfil():
    vehiculo = request.form.get("vehiculo", "moto")
    zona     = request.form.get("zona", "").strip()
    ciudad   = request.form.get("ciudad", "").strip()
    if ciudad not in get_ciudades_validas():
        cadete_actual = query("SELECT ciudad FROM cadetes WHERE usuario_id=?", (session["user_id"],), one=True)
        ciudad = (cadete_actual["ciudad"] if cadete_actual else None) or CIUDAD_DEFAULT
    execute("UPDATE cadetes SET vehiculo=?, zona=?, ciudad=? WHERE usuario_id=?",
            (vehiculo, zona, ciudad, session["user_id"]))
    flash("Perfil actualizado.", "success")
    return redirect(url_for("cadete_panel"))


# ── PANEL CLIENTE ─────────────────────────────────────────────────────────────

@app.route("/mi-cuenta")
@login_required
@rol_required("cliente")
def cliente_panel():
    pedidos = query("""
        SELECT p.*, r.nombre_local
        FROM pedidos p
        JOIN restaurantes r ON r.id = p.restaurante_id
        WHERE p.cliente_id = ?
        ORDER BY p.fecha_pedido DESC LIMIT 30
    """, (session["user_id"],))

    locales_frecuentes = query("""
        SELECT r.*, COUNT(p.id) as veces
        FROM pedidos p
        JOIN restaurantes r ON r.id = p.restaurante_id
        WHERE p.cliente_id = ? AND r.estado = 'aprobado'
        GROUP BY r.id
        ORDER BY veces DESC LIMIT 6
    """, (session["user_id"],))

    usuario = query("SELECT * FROM usuarios WHERE id = ?",
                    (session["user_id"],), one=True)
    cliente = query("SELECT * FROM clientes WHERE usuario_id = ?",
                    (session["user_id"],), one=True)

    favoritos = query("""
        SELECT r.* FROM favoritos f
        JOIN restaurantes r ON r.id = f.restaurante_id
        WHERE f.usuario_id = ? AND r.estado = 'aprobado'
        ORDER BY f.fecha DESC
    """, (session["user_id"],))

    return render_template("cliente_panel.html",
                           pedidos=pedidos,
                           locales_frecuentes=locales_frecuentes,
                           favoritos=favoritos,
                           usuario=usuario,
                           cliente=cliente)


@app.route("/favoritos/<int:restaurante_id>/toggle", methods=["POST"])
@login_required
@rol_required("cliente")
def favorito_toggle(restaurante_id):
    existe = query("SELECT id FROM favoritos WHERE usuario_id=? AND restaurante_id=?",
                    (session["user_id"], restaurante_id), one=True)
    if existe:
        execute("DELETE FROM favoritos WHERE id=?", (existe["id"],))
        return jsonify({"ok": True, "favorito": False})

    r = query("SELECT id FROM restaurantes WHERE id=? AND estado='aprobado'",
               (restaurante_id,), one=True)
    if not r:
        return jsonify({"error": "Local no encontrado"}), 404

    execute("INSERT INTO favoritos (usuario_id, restaurante_id) VALUES (?, ?)",
            (session["user_id"], restaurante_id))
    return jsonify({"ok": True, "favorito": True})


@app.route("/mi-cuenta/editar", methods=["POST"])
@login_required
@rol_required("cliente")
def cliente_editar_perfil():
    nombre    = request.form.get("nombre", "").strip()
    apellido  = request.form.get("apellido", "").strip()
    email     = request.form.get("email", "").strip().lower()
    telefono  = request.form.get("telefono", "").strip()
    dir_def   = request.form.get("direccion_default", "").strip()
    notas_def = request.form.get("notas_default", "").strip()

    execute("UPDATE usuarios SET nombre=?, apellido=?, email=?, telefono=? WHERE id=?",
            (nombre, apellido, email, telefono, session["user_id"]))
    execute("""
        UPDATE clientes SET direccion_default=?, notas_default=?
        WHERE usuario_id=?
    """, (dir_def, notas_def, session["user_id"]))

    session["nombre"] = nombre
    flash("Datos actualizados.", "success")
    return redirect(url_for("cliente_panel"))


@app.route("/mi-cuenta/password", methods=["GET", "POST"])
@login_required
@rol_required("cliente")
def cliente_cambiar_password():
    if request.method == "GET":
        flash("Usá el formulario para cambiar tu contraseña.", "info")
        return redirect(url_for("cliente_panel"))
    
    from werkzeug.security import check_password_hash, generate_password_hash
    actual = request.form.get("password_actual", "")
    nueva  = request.form.get("password_nueva", "")

    usuario = query("SELECT * FROM usuarios WHERE id = ?",
                    (session["user_id"],), one=True)

    if not check_password_hash(usuario["password_hash"], actual):
        flash("La contraseña actual es incorrecta.", "danger")
        return redirect(url_for("cliente_panel"))

    if len(nueva) < 6:
        flash("La nueva contraseña debe tener al menos 6 caracteres.", "danger")
        return redirect(url_for("cliente_panel"))

    execute("UPDATE usuarios SET password_hash=? WHERE id=?",
            (generate_password_hash(nueva), session["user_id"]))
    flash("Contraseña actualizada correctamente.", "success")
    return redirect(url_for("cliente_panel"))


# ── PANEL ADMIN ───────────────────────────────────────────────────────────────

@app.route("/admin")
@login_required
@rol_required("admin")
def admin_panel():
    pendientes_restaurantes = query("""
        SELECT r.*, u.nombre, u.apellido, u.email, u.telefono
        FROM restaurantes r JOIN usuarios u ON u.id = r.usuario_id
        WHERE r.estado = 'pendiente' ORDER BY r.fecha_alta
    """)
    pendientes_cadetes = query("""
        SELECT c.*, u.nombre, u.apellido, u.email, u.telefono
        FROM cadetes c JOIN usuarios u ON u.id = c.usuario_id
        WHERE c.estado = 'pendiente' ORDER BY u.fecha_registro
    """)
    todos_restaurantes = query("""
        SELECT r.*, u.nombre, u.apellido, u.email
        FROM restaurantes r JOIN usuarios u ON u.id = r.usuario_id
        ORDER BY r.estado, r.nombre_local
    """)
    todos_cadetes = query("""
        SELECT c.*, u.nombre, u.apellido, u.email, u.telefono
        FROM cadetes c JOIN usuarios u ON u.id = c.usuario_id
        ORDER BY c.estado, u.nombre
    """)
    ultimos_pedidos = query("""
        SELECT p.*, r.nombre_local,
               u.nombre AS nombre_cliente
        FROM pedidos p
        JOIN restaurantes r ON r.id = p.restaurante_id
        LEFT JOIN usuarios u ON u.id = p.cliente_id
        ORDER BY p.fecha_pedido DESC LIMIT 50
    """)
    stats = {
        "restaurantes_activos": query(
            "SELECT COUNT(*) as n FROM restaurantes WHERE estado='aprobado'", one=True)["n"],
        "cadetes_activos": query(
            "SELECT COUNT(*) as n FROM cadetes WHERE estado='aprobado'", one=True)["n"],
        "clientes": query(
            "SELECT COUNT(*) as n FROM usuarios WHERE rol='cliente'", one=True)["n"],
        "pedidos_hoy": query(
            "SELECT COUNT(*) as n FROM pedidos WHERE date(fecha_pedido)=CURRENT_DATE",
            one=True)["n"],
    }
    return render_template("admin_panel.html",
                           pendientes_restaurantes=pendientes_restaurantes,
                           pendientes_cadetes=pendientes_cadetes,
                           todos_restaurantes=todos_restaurantes,
                           todos_cadetes=todos_cadetes,
                           ultimos_pedidos=ultimos_pedidos,
                           stats=stats)


# ── ADMIN MÉTRICAS ────────────────────────────────────────────────────────────

@app.route("/admin/metricas")
@login_required
@rol_required("admin")
def admin_metricas():
    from datetime import datetime, timedelta
    hace_7_dias  = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')
    hace_30_dias = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')

    stats = {
        "usuarios_total":      query("SELECT COUNT(*) as n FROM usuarios", one=True)["n"],
        "clientes":            query("SELECT COUNT(*) as n FROM usuarios WHERE rol='cliente'", one=True)["n"],
        "restaurantes_activos":query("SELECT COUNT(*) as n FROM restaurantes WHERE estado='aprobado'", one=True)["n"],
        "cadetes_activos":     query("SELECT COUNT(*) as n FROM cadetes WHERE estado='aprobado'", one=True)["n"],
        "pedidos_total":       query("SELECT COUNT(*) as n FROM pedidos", one=True)["n"],
        "pedidos_hoy":         query("SELECT COUNT(*) as n FROM pedidos WHERE date(fecha_pedido)=CURRENT_DATE", one=True)["n"],
        "pedidos_semana":      query("SELECT COUNT(*) as n FROM pedidos WHERE fecha_pedido >= ?", (hace_7_dias,), one=True)["n"],
        "facturado_total":     query("SELECT COALESCE(SUM(total),0) as n FROM pedidos WHERE estado != 'cancelado'", one=True)["n"],
    }
    pedidos_por_dia = query("""
        SELECT date(fecha_pedido) as dia, COUNT(*) as total
        FROM pedidos WHERE fecha_pedido >= ?
        GROUP BY dia ORDER BY dia
    """, (hace_30_dias,))
    top_restaurantes = query("""
        SELECT r.nombre_local, COUNT(p.id) as pedidos,
               COALESCE(AVG(v.estrellas),0) as rating
        FROM restaurantes r
        LEFT JOIN pedidos p ON p.restaurante_id = r.id
        LEFT JOIN valoraciones v ON v.restaurante_id = r.id
        WHERE r.estado='aprobado'
        GROUP BY r.id ORDER BY pedidos DESC LIMIT 10
    """)
    return render_template("admin_metricas.html",
                           stats=stats,
                           pedidos_por_dia=pedidos_por_dia,
                           top_restaurantes=top_restaurantes)


@app.route("/admin/restaurante/<int:restaurante_id>/estado/<accion>", methods=["POST"])
@login_required
@rol_required("admin")
def admin_restaurante_estado(restaurante_id, accion):
    estados = {"aprobar": "aprobado", "suspender": "suspendido", "pendiente": "pendiente"}
    if accion not in estados:
        flash("Acción inválida.", "danger")
        return redirect(url_for("admin_panel"))
    execute("UPDATE restaurantes SET estado=? WHERE id=?",
            (estados[accion], restaurante_id))
    flash(f"Local {estados[accion]}.", "success")
    return redirect(url_for("admin_panel"))


@app.route("/admin/cadete/<int:cadete_id>/estado/<accion>", methods=["POST"])
@login_required
@rol_required("admin")
def admin_cadete_estado(cadete_id, accion):
    estados = {"aprobar": "aprobado", "suspender": "suspendido"}
    if accion not in estados:
        flash("Acción inválida.", "danger")
        return redirect(url_for("admin_panel"))
    execute("UPDATE cadetes SET estado=? WHERE id=?",
            (estados[accion], cadete_id))
    flash(f"Cadete {estados[accion]}.", "success")
    return redirect(url_for("admin_panel"))


# ── API: GENERAR LINK DE WHATSAPP ─────────────────────────────────────────────

import urllib.parse

def _limpiar_numero(tel):
    n = tel.replace("+","").replace("-","").replace(" ","")
    if not n.startswith("549"):
        n = "549" + n
    return n

def _wa_link(numero, mensaje):
    return f"https://wa.me/{numero}?text={urllib.parse.quote(mensaje)}"


@app.route("/api/whatsapp-link", methods=["POST"])
def whatsapp_link():
    data           = request.get_json()
    restaurante_id = data.get("restaurante_id")
    items          = data.get("items", [])
    tipo_entrega   = data.get("tipo_entrega", "retiro")
    direccion      = data.get("direccion", "").strip()
    notas          = data.get("notas", "").strip()
    nombre_cliente = data.get("nombre_cliente", "").strip()
    tel_cliente    = data.get("tel_cliente", "").strip()

    if not nombre_cliente:
        return jsonify({"error": "El nombre es obligatorio"}), 400
    if tipo_entrega == "delivery" and not direccion:
        return jsonify({"error": "La dirección es obligatoria para delivery"}), 400

    restaurante = query(
        "SELECT * FROM restaurantes WHERE id = ? AND estado = 'aprobado'",
        (restaurante_id,), one=True
    )
    if not restaurante or not items:
        return jsonify({"error": "Datos inválidos"}), 400

    def _sub_item(i):
        return i["cantidad"] * i["precio"] * (i.get("peso_kg") or 1)

    subtotal = sum(_sub_item(i) for i in items)

    # El descuento por promo general del local se recalcula acá (no se
    # confía en lo que mande el navegador) para que no se pueda inflar
    # el descuento manipulando la petición.
    promo_orden = query("""
        SELECT * FROM promociones
        WHERE restaurante_id = ? AND activa = 1 AND producto_id IS NULL
          AND (fecha_inicio IS NULL OR fecha_inicio <= CURRENT_DATE)
          AND (fecha_fin   IS NULL OR fecha_fin   >= CURRENT_DATE)
        ORDER BY fecha_creacion DESC LIMIT 1
    """, (restaurante_id,), one=True)

    descuento = 0
    if promo_orden:
        if promo_orden["tipo_descuento"] == "porcentaje" and promo_orden["descuento_pct"]:
            descuento = round(subtotal * promo_orden["descuento_pct"] / 100)
        elif promo_orden["tipo_descuento"] == "monto" and promo_orden["descuento_monto"]:
            descuento = min(subtotal, promo_orden["descuento_monto"])

    # El costo de envío se toma del local (restaurantes.costo_envio) y se
    # suma al total, y también se guarda "congelado" en el pedido (columna
    # pedidos.costo_envio) para que quede constante aunque el local cambie
    # después su tarifa — así el cadete siempre puede verificar cuánto le
    # corresponde reclamar por ESE pedido puntual.
    envio = 0
    if tipo_entrega == "delivery" and restaurante["hace_envio"]:
        envio = restaurante["costo_envio"] or 0

    total = subtotal - descuento + envio

    def _pesos(n):
        # Formatea como "$1.234" (separador de miles con punto), sin tocar
        # ningún otro texto de la línea (antes se usaba .replace(",",".")
        # sobre la línea entera, lo que también rompía comas dentro de
        # nombres de producto, p.ej. listas de gustos "Choco, Vainilla").
        return f"${n:,.0f}".replace(",", ".")

    lineas = [f"*Nuevo pedido — {restaurante['nombre_local']}*\n"]
    lineas.append(f"*Cliente:* {nombre_cliente}")
    if tel_cliente:
        lineas.append(f"*Tel:* {tel_cliente}")
    lineas.append("")
    for item in items:
        sub = _sub_item(item)
        # El nombre del item ya incluye el peso y los gustos elegidos (armado
        # en el frontend), así que no hace falta un formato especial acá.
        lineas.append(f"• {item['cantidad']}x {item['nombre']} — {_pesos(sub)}")
    if descuento > 0:
        lineas.append(f"\n*Subtotal: {_pesos(subtotal)}*")
        lineas.append(f"*{promo_orden['titulo']}: -{_pesos(descuento)}*")
    if envio > 0:
        lineas.append(f"*Envío: {_pesos(envio)}*")
    lineas.append(f"\n*Total: {_pesos(total)}*")
    lineas.append(f"*Entrega:* {'Delivery' if tipo_entrega == 'delivery' else 'Retiro en local'}")
    if tipo_entrega == "delivery":
        lineas.append(f"*Dirección:* {direccion}")
    if notas:
        lineas.append(f"*Notas:* {notas}")
    lineas.append("\n_Pedido generado desde pediaca.ar_")

    mensaje    = "\n".join(lineas)
    numero     = _limpiar_numero(restaurante["whatsapp"])
    link       = _wa_link(numero, mensaje)

    cliente_id = session.get("user_id")
    nom_anon   = nombre_cliente if not cliente_id else None
    tel_anon   = tel_cliente    if not cliente_id else None

    # Código de 4 dígitos que el cliente le va a dar al cadete al recibir el
    # pedido, para que el cadete confirme la entrega correcta (ver
    # /cadete/pedido/<id>/entregar). Sólo aplica a delivery, pero se genera
    # siempre por simplicidad — en retiro nadie lo usa.
    codigo_entrega = f"{random.randint(0, 9999):04d}"

    pedido_id = execute("""
        INSERT INTO pedidos
            (restaurante_id, cliente_id, nombre_cliente_anonimo, telefono_cliente_anonimo,
             tipo_entrega, direccion_entrega, total, costo_envio, notas, codigo_entrega, enviado_whatsapp)
        VALUES (?,?,?,?,?,?,?,?,?,?,1)
    """, (restaurante_id, cliente_id, nom_anon, tel_anon,
          tipo_entrega, direccion, total, envio, notas, codigo_entrega))

    for item in items:
        execute("""
            INSERT INTO items_pedido
                (pedido_id, producto_id, nombre_producto, cantidad, precio_unitario, subtotal, peso_kg)
            VALUES (?,?,?,?,?,?,?)
        """, (pedido_id, item.get("producto_id"), item["nombre"],
              item["cantidad"], item["precio"], _sub_item(item), item.get("peso_kg")))

    return jsonify({"link": link, "pedido_id": pedido_id})


# ── SEGUIMIENTO DE PEDIDO ─────────────────────────────────────────────────────

@app.route("/pedido/<int:pedido_id>")
def seguimiento_pedido(pedido_id):
    # Página pública (hay pedidos de clientes anónimos) — por eso no expone
    # dirección de entrega ni teléfonos, sólo lo necesario para trackear.
    pedido = query("""
        SELECT p.id, p.estado, p.tipo_entrega, p.total, p.fecha_pedido, p.codigo_entrega,
               r.nombre_local, r.logo_url,
               u.nombre AS cadete_nombre, c.vehiculo AS cadete_vehiculo
        FROM pedidos p
        JOIN restaurantes r ON r.id = p.restaurante_id
        LEFT JOIN cadetes c ON c.id = p.cadete_id
        LEFT JOIN usuarios u ON u.id = c.usuario_id
        WHERE p.id = ?
    """, (pedido_id,), one=True)
    if not pedido:
        flash("Pedido no encontrado.", "danger")
        return redirect(url_for("home"))

    items = query("""
        SELECT nombre_producto, cantidad, subtotal
        FROM items_pedido WHERE pedido_id = ?
        ORDER BY id
    """, (pedido_id,))

    return render_template("seguimiento_pedido.html", pedido=pedido, items=items)


# ── API: STATUS DEL PEDIDO ───────────────────────────────────────────────────

@app.route("/api/pedido/<int:pedido_id>/status")
def pedido_status(pedido_id):
    # Endpoint público (hay pedidos de clientes anónimos, sin login) —
    # por eso no exige @login_required. Para no filtrar datos personales
    # del cadete a cualquiera que adivine un id, nunca se devuelve su
    # teléfono acá, sólo nombre/vehículo.
    pedido = query("""
        SELECT p.estado, p.cadete_id,
               u.nombre AS cadete_nombre,
               c.vehiculo
        FROM pedidos p
        LEFT JOIN cadetes c ON c.id = p.cadete_id
        LEFT JOIN usuarios u ON u.id = c.usuario_id
        WHERE p.id = ?
    """, (pedido_id,), one=True)
    if not pedido:
        return jsonify({"error": "No encontrado"}), 404
    return jsonify({
        "estado":          pedido["estado"],
        "cadete_nombre":   pedido["cadete_nombre"],
        "cadete_vehiculo": pedido["vehiculo"],
    })


# ── API: UBICACIÓN GPS EN VIVO DEL CADETE ─────────────────────────────────────

@app.route("/api/cadete/ubicacion", methods=["POST"])
@login_required
@rol_required("cadete")
def cadete_actualizar_ubicacion():
    """El celular del cadete manda su posición acá mientras tiene una entrega
    en curso (ver JS en cadete_panel.html). Sólo puede actualizar su propia
    fila — se identifica por la sesión, nunca por un id que venga del cliente."""
    data = request.get_json(silent=True) or {}
    try:
        lat = float(data.get("lat"))
        lng = float(data.get("lng"))
    except (TypeError, ValueError):
        return jsonify({"error": "Coordenadas inválidas"}), 400
    execute("""
        UPDATE cadetes SET lat=?, lng=?, ubicacion_ts=CURRENT_TIMESTAMP
        WHERE usuario_id=?
    """, (lat, lng, session["user_id"]))
    return jsonify({"ok": True})


@app.route("/api/pedido/<int:pedido_id>/tracking")
def pedido_tracking(pedido_id):
    """Devuelve las coordenadas para pintar el mapa de seguimiento: ubicación
    del local, del destino de entrega (si es delivery) y del cadete en vivo
    (si el pedido está en camino). Público como /status — hay pedidos de
    clientes anónimos que necesitan ver esto sin login. Nunca se expone acá
    ni teléfono ni nombre del cadete (eso ya lo maneja /status)."""
    pedido = query("""
        SELECT p.id, p.estado, p.tipo_entrega, p.direccion_entrega,
               p.lat_entrega, p.lng_entrega, p.cadete_id, p.restaurante_id,
               r.direccion AS restaurante_direccion, r.ciudad,
               r.lat AS restaurante_lat, r.lng AS restaurante_lng,
               c.lat AS cadete_lat, c.lng AS cadete_lng, c.ubicacion_ts
        FROM pedidos p
        JOIN restaurantes r ON r.id = p.restaurante_id
        LEFT JOIN cadetes c ON c.id = p.cadete_id
        WHERE p.id = ?
    """, (pedido_id,), one=True)
    if not pedido:
        return jsonify({"error": "No encontrado"}), 404

    r_lat, r_lng = pedido["restaurante_lat"], pedido["restaurante_lng"]
    if r_lat is None and pedido["restaurante_direccion"]:
        coords = geocodificar_direccion(pedido["restaurante_direccion"], pedido["ciudad"])
        if coords:
            r_lat, r_lng = coords
            execute("UPDATE restaurantes SET lat=?, lng=? WHERE id=?",
                    (r_lat, r_lng, pedido["restaurante_id"]))

    e_lat, e_lng = pedido["lat_entrega"], pedido["lng_entrega"]
    if pedido["tipo_entrega"] == "delivery" and e_lat is None and pedido["direccion_entrega"]:
        coords = geocodificar_direccion(pedido["direccion_entrega"], pedido["ciudad"])
        if coords:
            e_lat, e_lng = coords
            execute("UPDATE pedidos SET lat_entrega=?, lng_entrega=? WHERE id=?",
                    (e_lat, e_lng, pedido_id))

    cadete_pos = None
    if pedido["estado"] == "en_camino" and pedido["cadete_lat"] is not None:
        cadete_pos = {
            "lat": pedido["cadete_lat"],
            "lng": pedido["cadete_lng"],
            "actualizado": pedido["ubicacion_ts"].isoformat() if hasattr(pedido["ubicacion_ts"], "isoformat") else pedido["ubicacion_ts"],
        }

    return jsonify({
        "estado":       pedido["estado"],
        "tipo_entrega": pedido["tipo_entrega"],
        "restaurante":  {"lat": r_lat, "lng": r_lng} if r_lat is not None else None,
        "entrega":      {"lat": e_lat, "lng": e_lng} if e_lat is not None else None,
        "cadete":       cadete_pos,
    })


# ── RESTAURANTE: MÉTRICAS PROPIAS ─────────────────────────────────────────────

@app.route("/mi-local/metricas")
@login_required
@rol_required("restaurante")
def restaurante_metricas():
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        flash("Tu local debe estar aprobado para ver métricas.", "warning")
        return redirect(url_for("restaurante_panel"))

    from datetime import datetime, timedelta
    hace_30_dias = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')
    rid = restaurante["id"]

    stats = {
        "pedidos_total": query(
            "SELECT COUNT(*) as n FROM pedidos WHERE restaurante_id=?", (rid,), one=True)["n"],
        "pedidos_hoy": query(
            "SELECT COUNT(*) as n FROM pedidos WHERE restaurante_id=? AND date(fecha_pedido)=CURRENT_DATE",
            (rid,), one=True)["n"],
        "facturado_mes": query("""
            SELECT COALESCE(SUM(total),0) as n FROM pedidos
            WHERE restaurante_id=? AND estado != 'cancelado' AND fecha_pedido >= ?
        """, (rid, hace_30_dias), one=True)["n"],
        "rating_promedio": query(
            "SELECT COALESCE(AVG(estrellas),0) as n FROM valoraciones WHERE restaurante_id=?",
            (rid,), one=True)["n"],
    }

    productos_top = query("""
        SELECT ip.nombre_producto, SUM(ip.cantidad) as cantidad_total, SUM(ip.subtotal) as ingresos
        FROM items_pedido ip
        JOIN pedidos p ON p.id = ip.pedido_id
        WHERE p.restaurante_id = ? AND p.estado != 'cancelado'
        GROUP BY ip.nombre_producto
        ORDER BY cantidad_total DESC LIMIT 8
    """, (rid,))

    pedidos_por_dia = query("""
        SELECT date(fecha_pedido) as dia, COUNT(*) as total
        FROM pedidos WHERE restaurante_id = ? AND fecha_pedido >= ?
        GROUP BY dia ORDER BY dia
    """, (rid, hace_30_dias))

    return render_template("restaurante_metricas.html",
                           restaurante=restaurante,
                           stats=stats,
                           productos_top=productos_top,
                           pedidos_por_dia=pedidos_por_dia)


# ── RESTAURANTE: HISTORIAL DE PEDIDOS ─────────────────────────────────────────

@app.route("/mi-local/pedidos")
@login_required
@rol_required("restaurante")
def restaurante_pedidos():
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        return redirect(url_for("restaurante_panel"))

    pedidos = query("""
        SELECT p.*,
               u.nombre  AS nombre_cliente,
               u.telefono AS cliente_tel,
               uc.nombre  AS cadete_nombre,
               uc.telefono AS cadete_tel,
               c.vehiculo  AS cadete_vehiculo
        FROM pedidos p
        LEFT JOIN usuarios u  ON u.id  = p.cliente_id
        LEFT JOIN cadetes  c  ON c.id  = p.cadete_id
        LEFT JOIN usuarios uc ON uc.id = c.usuario_id
        WHERE p.restaurante_id = ?
        ORDER BY p.fecha_pedido DESC
        LIMIT 100
    """, (restaurante["id"],))

    items_por_pedido = {}
    if pedidos:
        ids = ",".join(str(p["id"]) for p in pedidos)
        items = query(f"""
            SELECT * FROM items_pedido WHERE pedido_id IN ({ids})
            ORDER BY pedido_id, id
        """)
        for it in items:
            items_por_pedido.setdefault(it["pedido_id"], []).append(it)

    return render_template("restaurante_pedidos.html",
                           restaurante=restaurante,
                           pedidos=pedidos,
                           items_por_pedido=items_por_pedido)


@app.route("/mi-local/pedido/<int:pedido_id>/estado/<nuevo_estado>", methods=["POST"])
@login_required
@rol_required("restaurante")
def restaurante_cambiar_estado(pedido_id, nuevo_estado):
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        return redirect(url_for("restaurante_panel"))

    estados_validos = ["confirmado", "cancelado", "entregado"]
    if nuevo_estado not in estados_validos:
        flash("Estado inválido.", "danger")
        return redirect(url_for("restaurante_pedidos"))

    execute("""
        UPDATE pedidos SET estado=?, fecha_actualizado=CURRENT_TIMESTAMP
        WHERE id=? AND restaurante_id=?
    """, (nuevo_estado, pedido_id, restaurante["id"]))

    if nuevo_estado == "confirmado":
        pedido = query("SELECT * FROM pedidos WHERE id=? AND restaurante_id=?",
                        (pedido_id, restaurante["id"]), one=True)
        if pedido and pedido["tipo_entrega"] == "delivery":
            notificar_cadetes_push(
                pedido_id,
                restaurante["nombre_local"],
                pedido["total"],
                pedido["direccion_entrega"]
            )

    if nuevo_estado == "cancelado":
        pedido = query("SELECT cadete_id FROM pedidos WHERE id=? AND restaurante_id=?",
                        (pedido_id, restaurante["id"]), one=True)
        if pedido and pedido["cadete_id"]:
            notificar_cadete_pedido_cancelado(pedido_id, pedido["cadete_id"], restaurante["nombre_local"])

    notificar_cliente_push(pedido_id, nuevo_estado, restaurante["nombre_local"])

    flash(f"Pedido #{pedido_id} marcado como {nuevo_estado}.", "success")
    return redirect(url_for("restaurante_pedidos"))


@app.route("/api/notificar-cadetes/<int:pedido_id>")
@login_required
@rol_required("restaurante")
def notificar_cadetes(pedido_id):
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        return jsonify({"error": "No autorizado"}), 403

    pedido = query(
        "SELECT * FROM pedidos WHERE id=? AND restaurante_id=?",
        (pedido_id, restaurante["id"]), one=True
    )
    if not pedido:
        return jsonify({"error": "Pedido no encontrado"}), 404

    cadetes = query("""
        SELECT u.nombre, u.telefono, c.vehiculo, c.zona
        FROM cadetes c
        JOIN usuarios u ON u.id = c.usuario_id
        WHERE c.estado='aprobado' AND c.disponible=1 AND u.telefono IS NOT NULL
    """)

    mensaje = (
        f"*PediAcá — Pedido disponible!*\n\n"
        f"Local: {restaurante['nombre_local']}\n"
        f"Retiro: {restaurante['direccion'] or 'A confirmar con el local'}\n"
        f"Entrega: {pedido['direccion_entrega'] or 'Retiro en local'}\n"
        f"Total del pedido: ${int(pedido['total'])}\n\n"
        f"Entrá a pediaca.ar para aceptarlo antes que otro cadete."
    )

    links = []
    for c in cadetes:
        if c["telefono"]:
            num  = _limpiar_numero(c["telefono"])
            links.append({
                "nombre":  c["nombre"],
                "vehiculo":c["vehiculo"],
                "link":    _wa_link(num, mensaje)
            })

    return jsonify({"cadetes": links, "pedido_id": pedido_id})


@app.route("/pedido/<int:pedido_id>/en-camino", methods=["POST"])
@login_required
def marcar_en_camino(pedido_id):
    rol = session.get("rol")

    if rol == "restaurante":
        restaurante = get_restaurante_aprobado()
        if not restaurante:
            return jsonify({"error": "No autorizado"}), 403
        pedido = query("SELECT * FROM pedidos WHERE id=? AND restaurante_id=?",
                       (pedido_id, restaurante["id"]), one=True)
    elif rol == "cadete":
        cadete = query("SELECT * FROM cadetes WHERE usuario_id=? AND estado='aprobado'",
                       (session["user_id"],), one=True)
        if not cadete:
            return jsonify({"error": "No autorizado"}), 403
        pedido = query("SELECT * FROM pedidos WHERE id=? AND cadete_id=?",
                       (pedido_id, cadete["id"]), one=True)
    else:
        return jsonify({"error": "No autorizado"}), 403

    if not pedido:
        return jsonify({"error": "Pedido no encontrado"}), 404

    execute("""
        UPDATE pedidos SET estado='en_camino', fecha_actualizado=CURRENT_TIMESTAMP
        WHERE id=?
    """, (pedido_id,))

    local = query("SELECT nombre_local FROM restaurantes WHERE id=?", (pedido["restaurante_id"],), one=True)
    notificar_cliente_push(pedido_id, "en_camino", local["nombre_local"] if local else "tu local")

    return jsonify({"ok": True})


# ── CADETE: ACEPTAR PEDIDO ────────────────────────────────────────────────────

@app.route("/cadete/aceptar/<int:pedido_id>", methods=["POST"])
@login_required
@rol_required("cadete")
def cadete_aceptar_pedido(pedido_id):
    cadete = query(
        "SELECT * FROM cadetes WHERE usuario_id=? AND estado='aprobado'",
        (session["user_id"],), one=True
    )
    if not cadete:
        return jsonify({"error": "No autorizado"}), 403

    pedido = query(
        "SELECT * FROM pedidos WHERE id=? AND cadete_id IS NULL AND estado='confirmado'",
        (pedido_id,), one=True
    )
    if not pedido:
        return jsonify({"ok": False, "msg": "El pedido ya fue tomado por otro cadete"}), 409

    execute("""
        UPDATE pedidos SET cadete_id=?, estado='en_camino',
               fecha_actualizado=CURRENT_TIMESTAMP
        WHERE id=? AND cadete_id IS NULL
    """, (cadete["id"], pedido_id))

    check = query("SELECT cadete_id FROM pedidos WHERE id=?", (pedido_id,), one=True)
    if check["cadete_id"] != cadete["id"]:
        return jsonify({"ok": False, "msg": "El pedido ya fue tomado por otro cadete"}), 409

    local = query("SELECT nombre_local FROM restaurantes WHERE id=?", (pedido["restaurante_id"],), one=True)
    notificar_cliente_push(pedido_id, "en_camino", local["nombre_local"] if local else "tu local")

    return jsonify({"ok": True, "msg": "¡Pedido aceptado! Coordiná con el local."})


@app.route("/api/pedidos-nuevos")
@login_required
@rol_required("cadete")
def pedidos_nuevos_cadete():
    cadete = query("SELECT ciudad FROM cadetes WHERE usuario_id = ?", (session["user_id"],), one=True)
    ciudad_cadete = (cadete["ciudad"] if cadete else None) or CIUDAD_DEFAULT
    pedidos = query("""
        SELECT p.id, p.total, p.costo_envio, p.direccion_entrega, p.fecha_pedido,
               r.nombre_local, r.direccion AS local_direccion, r.whatsapp
        FROM pedidos p
        JOIN restaurantes r ON r.id = p.restaurante_id
        WHERE p.tipo_entrega='delivery'
          AND p.estado='confirmado'
          AND p.cadete_id IS NULL
          AND r.ciudad = ?
        ORDER BY p.fecha_pedido DESC
    """, (ciudad_cadete,))
    return jsonify({"pedidos": [dict(p) for p in pedidos]})


# ── FLYER ─────────────────────────────────────────────────────────────────────

@app.route("/mi-local/flyer")
@login_required
@rol_required("restaurante")
def descargar_flyer():
    from flyer import generar_flyer
    from flask import send_file
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        flash("Tu local debe estar aprobado para descargar el flyer.", "warning")
        return redirect(url_for("restaurante_panel"))

    logo_path = None
    if restaurante["logo_url"]:
        # logo_url puede ser una URL de Cloudinary (http...) o una ruta
        # relativa dentro de static/ (almacenamiento local) — flyer.py
        # sabe manejar ambos casos.
        if restaurante["logo_url"].startswith("http"):
            logo_path = restaurante["logo_url"]
        else:
            logo_path = os.path.join("static", restaurante["logo_url"])

    png_bytes = generar_flyer(
        nombre_local   = restaurante["nombre_local"],
        restaurante_id = restaurante["id"],
        logo_path      = logo_path,
        categoria      = restaurante["categoria"] or "",
    )

    nombre_archivo = f"flyer_pediaca_{restaurante['nombre_local'].replace(' ','_').lower()}.png"
    return send_file(
        io.BytesIO(png_bytes),
        mimetype="image/png",
        as_attachment=True,
        download_name=nombre_archivo
    )


# ── UPLOAD FOTOS RESTAURANTE ──────────────────────────────────────────────────

@app.route("/mi-local/foto/logo", methods=["POST"])
@login_required
@rol_required("restaurante")
def subir_logo():
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        return redirect(url_for("restaurante_panel"))
    archivo = request.files.get("logo")
    ruta    = guardar_imagen(archivo, "logos")
    if ruta:
        execute("UPDATE restaurantes SET logo_url=? WHERE id=?",
                (ruta, restaurante["id"]))
        flash("Logo actualizado.", "success")
    else:
        flash("Archivo inválido. Usá PNG, JPG o WEBP.", "danger")
    return redirect(url_for("restaurante_panel"))


@app.route("/mi-local/foto/banner", methods=["POST"])
@login_required
@rol_required("restaurante")
def subir_banner():
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        return redirect(url_for("restaurante_panel"))
    archivo = request.files.get("banner")
    ruta    = guardar_imagen(archivo, "banners")
    if ruta:
        execute("UPDATE restaurantes SET banner_url=? WHERE id=?",
                (ruta, restaurante["id"]))
        flash("Banner actualizado.", "success")
    else:
        flash("Archivo inválido. Usá PNG, JPG o WEBP.", "danger")
    return redirect(url_for("restaurante_panel"))


@app.route("/mi-local/producto/<int:prod_id>/foto", methods=["POST"])
@login_required
@rol_required("restaurante")
def subir_foto_producto(prod_id):
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        return redirect(url_for("restaurante_panel"))
    archivo = request.files.get("foto")
    ruta    = guardar_imagen(archivo, "productos")
    if ruta:
        execute("UPDATE productos SET foto_url=? WHERE id=? AND restaurante_id=?",
                (ruta, prod_id, restaurante["id"]))
        flash("Foto del producto actualizada.", "success")
    else:
        flash("Archivo inválido.", "danger")
    return redirect(url_for("restaurante_panel"))


# ── PROMOCIONES ───────────────────────────────────────────────────────────────

@app.route("/mi-local/promocion/nueva", methods=["POST"])
@login_required
@rol_required("restaurante")
def promocion_nueva():
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        return redirect(url_for("restaurante_panel"))
    
    titulo = request.form.get("titulo", "").strip()
    descripcion = request.form.get("descripcion", "").strip()
    tipo_descuento = request.form.get("tipo_descuento", "porcentaje")
    fecha_inicio = request.form.get("fecha_inicio") or None
    fecha_fin = request.form.get("fecha_fin") or None
    producto_id = request.form.get("producto_id") or None
    archivo = request.files.get("imagen")
    imagen_url = guardar_imagen(archivo, "promociones") if archivo and archivo.filename else None

    if not titulo:
        flash("El título es obligatorio.", "danger")
        return redirect(url_for("restaurante_panel") + "#sec-promociones")

    try:
        descuento_pct = int(request.form.get("descuento_pct", 0) or 0)
        descuento_monto = int(request.form.get("descuento_monto", 0) or 0)
        precio_con_descuento = request.form.get("precio_con_descuento") or None
        if precio_con_descuento:
            precio_con_descuento = int(precio_con_descuento)
    except ValueError:
        flash("Los valores de descuento/precio tienen que ser números.", "danger")
        return redirect(url_for("restaurante_panel") + "#sec-promociones")

    execute("""
        INSERT INTO promociones
            (restaurante_id, titulo, descripcion, imagen_url, tipo_descuento,
             descuento_pct, descuento_monto, fecha_inicio, fecha_fin, activa,
             producto_id, precio_con_descuento)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
    """, (restaurante["id"], titulo, descripcion, imagen_url, tipo_descuento,
          descuento_pct, descuento_monto, fecha_inicio, fecha_fin,
          producto_id, precio_con_descuento))

    flash(f"Promoción '{titulo}' creada.", "success")
    return redirect(url_for("restaurante_panel") + "#sec-promociones")

@app.route("/mi-local/promocion/<int:promo_id>/toggle", methods=["POST"])
@login_required
@rol_required("restaurante")
def promocion_toggle(promo_id):
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        return redirect(url_for("restaurante_panel"))
    execute("""
        UPDATE promociones SET activa = 1 - activa
        WHERE id = ? AND restaurante_id = ?
    """, (promo_id, restaurante["id"]))
    return redirect(url_for("restaurante_panel") + "#sec-promociones")

@app.route("/mi-local/promocion/<int:promo_id>/eliminar", methods=["POST"])
@login_required
@rol_required("restaurante")
def promocion_eliminar(promo_id):
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        return redirect(url_for("restaurante_panel"))
    execute("DELETE FROM promociones WHERE id=? AND restaurante_id=?",
            (promo_id, restaurante["id"]))
    flash("Promoción eliminada.", "success")
    return redirect(url_for("restaurante_panel") + "#sec-promociones")

@app.route("/mi-local/promocion/<int:promo_id>/editar", methods=["POST"])
@login_required
@rol_required("restaurante")
def promocion_editar(promo_id):
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        return redirect(url_for("restaurante_panel"))
    
    titulo = request.form.get("titulo", "").strip()
    descripcion = request.form.get("descripcion", "").strip()
    tipo_descuento = request.form.get("tipo_descuento", "porcentaje")
    fecha_inicio = request.form.get("fecha_inicio") or None
    fecha_fin = request.form.get("fecha_fin") or None
    archivo = request.files.get("imagen")

    try:
        descuento_pct = int(request.form.get("descuento_pct", 0) or 0)
        descuento_monto = int(request.form.get("descuento_monto", 0) or 0)
    except ValueError:
        flash("Los valores de descuento tienen que ser números.", "danger")
        return redirect(url_for("restaurante_panel") + "#sec-promociones")

    promo = query("SELECT * FROM promociones WHERE id=? AND restaurante_id=?",
                  (promo_id, restaurante["id"]), one=True)
    if not promo:
        return redirect(url_for("restaurante_panel"))
    
    nueva_imagen = guardar_imagen(archivo, "promociones") if archivo and archivo.filename else promo["imagen_url"]
    
    execute("""
        UPDATE promociones SET
            titulo=?, descripcion=?, imagen_url=?,
            tipo_descuento=?, descuento_pct=?, descuento_monto=?,
            fecha_inicio=?, fecha_fin=?
        WHERE id=? AND restaurante_id=?
    """, (titulo, descripcion, nueva_imagen, tipo_descuento,
          descuento_pct, descuento_monto, fecha_inicio, fecha_fin,
          promo_id, restaurante["id"]))
    
    flash("Promoción actualizada.", "success")
    return redirect(url_for("restaurante_panel") + "#sec-promociones")


# ── TOGGLE ABIERTO/CERRADO ────────────────────────────────────────────────────

@app.route("/mi-local/toggle-abierto", methods=["POST"])
@login_required
@rol_required("restaurante")
def restaurante_toggle_abierto():
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        return redirect(url_for("restaurante_panel"))
    execute("UPDATE restaurantes SET abierto = 1 - abierto WHERE id=?",
            (restaurante["id"],))
    return redirect(url_for("restaurante_panel"))


# ── RECUPERO DE CONTRASEÑA VIA WHATSAPP ───────────────────────────────────────

@app.route("/recuperar-password", methods=["GET", "POST"])
def recuperar_password():
    if request.method == "POST":
        telefono = request.form.get("telefono", "").strip()
        email    = request.form.get("email", "").strip().lower()

        usuario = None
        if telefono:
            usuario = query("SELECT * FROM usuarios WHERE telefono=?", (telefono,), one=True)
        if not usuario and email:
            usuario = query("SELECT * FROM usuarios WHERE email=?", (email,), one=True)

        if usuario and usuario["telefono"]:
            import secrets
            from datetime import datetime, timedelta
            token  = secrets.token_urlsafe(32)
            expira = (datetime.now() + timedelta(hours=2)).strftime("%Y-%m-%d %H:%M:%S")
            execute("""
                INSERT INTO password_reset_tokens (usuario_id, token, expira)
                VALUES (?,?,?)
            """, (usuario["id"], token, expira))

            base   = request.host_url.rstrip("/")
            link   = f"{base}/reset-password/{token}"
            msg    = f"PediAcá · Hola {usuario['nombre']}! Para resetear tu contraseña entrá a este link (válido 2hs): {link}"
            numero = _limpiar_numero(usuario["telefono"])
            wa_link = _wa_link(numero, msg)
            return render_template("recuperar_password.html",
                                   wa_link=wa_link, usuario=usuario, enviado=True)
        else:
            flash("No encontramos ninguna cuenta con esos datos.", "danger")

    return render_template("recuperar_password.html", enviado=False)


@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    from datetime import datetime
    ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    registro = query("""
        SELECT t.*, u.nombre FROM password_reset_tokens t
        JOIN usuarios u ON u.id = t.usuario_id
        WHERE t.token=? AND t.usado=0 AND t.expira > ?
    """, (token, ahora), one=True)

    if not registro:
        flash("El link expiró o ya fue usado.", "danger")
        return redirect(url_for("login"))

    if request.method == "POST":
        from werkzeug.security import generate_password_hash
        nueva  = request.form.get("password", "")
        nueva2 = request.form.get("password2", "")
        if len(nueva) < 6:
            flash("La contraseña debe tener al menos 6 caracteres.", "danger")
            return render_template("reset_password.html", token=token, nombre=registro["nombre"])
        if nueva != nueva2:
            flash("Las contraseñas no coinciden.", "danger")
            return render_template("reset_password.html", token=token, nombre=registro["nombre"])

        execute("UPDATE usuarios SET password_hash=? WHERE id=?",
                (generate_password_hash(nueva), registro["usuario_id"]))
        execute("UPDATE password_reset_tokens SET usado=1 WHERE token=?", (token,))
        flash("¡Contraseña cambiada! Ya podés iniciar sesión.", "success")
        return redirect(url_for("login"))

    return render_template("reset_password.html", token=token, nombre=registro["nombre"])


# ── VALORACIONES ──────────────────────────────────────────────────────────────

@app.route("/valorar/<int:pedido_id>", methods=["POST"])
@login_required
def valorar_pedido(pedido_id):
    try:
        estrellas = int(request.form.get("estrellas", 5))
    except ValueError:
        estrellas = 0
    if not 1 <= estrellas <= 5:
        flash("La valoración tiene que ser entre 1 y 5 estrellas.", "danger")
        return redirect(url_for("cliente_panel"))
    comentario = request.form.get("comentario", "").strip()

    pedido = query("SELECT * FROM pedidos WHERE id=? AND cliente_id=? AND estado='entregado'",
                   (pedido_id, session["user_id"]), one=True)
    if not pedido:
        flash("No podés valorar este pedido.", "danger")
        return redirect(url_for("cliente_panel"))

    ya_valorado = query("SELECT id FROM valoraciones WHERE pedido_id=?", (pedido_id,), one=True)
    if ya_valorado:
        flash("Ya valoraste este pedido.", "warning")
        return redirect(url_for("cliente_panel"))

    execute("""
        INSERT INTO valoraciones (pedido_id, restaurante_id, cliente_id, estrellas, comentario)
        VALUES (?,?,?,?,?)
    """, (pedido_id, pedido["restaurante_id"], session["user_id"], estrellas, comentario))
    flash("¡Gracias por tu valoración!", "success")
    return redirect(url_for("cliente_panel"))


# ── CADETE: RECHAZAR PEDIDO ───────────────────────────────────────────────────

@app.route("/cadete/rechazar/<int:pedido_id>", methods=["POST"])
@login_required
@rol_required("cadete")
def cadete_rechazar_pedido(pedido_id):
    cadete = query("SELECT * FROM cadetes WHERE usuario_id=? AND estado='aprobado'",
                   (session["user_id"],), one=True)
    if not cadete:
        return jsonify({"error": "No autorizado"}), 403
    execute("""
        UPDATE pedidos SET cadete_id=NULL, estado='confirmado',
               fecha_actualizado=CURRENT_TIMESTAMP
        WHERE id=? AND cadete_id=?
    """, (pedido_id, cadete["id"]))
    return jsonify({"ok": True})


# ── CADETE: MARCAR ENTREGADO (con código de seguridad) ───────────────────────

@app.route("/cadete/pedido/<int:pedido_id>/entregar", methods=["POST"])
@login_required
@rol_required("cadete")
def cadete_marcar_entregado(pedido_id):
    """El cadete confirma la entrega ingresando el código de 4 dígitos que
    le mostró el cliente (pantalla de seguimiento). Evita el caso de tocar
    'entregado' por error o entregarle el pedido a la persona equivocada.
    Pedidos viejos (creados antes de este feature) no tienen código
    guardado — para esos se permite confirmar sin pedirlo."""
    cadete = query("SELECT * FROM cadetes WHERE usuario_id=? AND estado='aprobado'",
                   (session["user_id"],), one=True)
    if not cadete:
        return jsonify({"ok": False, "msg": "No autorizado"}), 403

    pedido = query("""
        SELECT p.*, r.nombre_local FROM pedidos p
        JOIN restaurantes r ON r.id = p.restaurante_id
        WHERE p.id=? AND p.cadete_id=?
    """, (pedido_id, cadete["id"]), one=True)
    if not pedido:
        return jsonify({"ok": False, "msg": "Pedido no encontrado"}), 404
    if pedido["estado"] != "en_camino":
        return jsonify({"ok": False, "msg": "Este pedido no está en camino."}), 400

    codigo_ingresado = (request.get_json(silent=True) or {}).get("codigo", "").strip()
    if pedido["codigo_entrega"]:
        if codigo_ingresado != pedido["codigo_entrega"]:
            return jsonify({"ok": False, "msg": "Código incorrecto. Pedile al cliente que te muestre el código de 4 dígitos de su pantalla de seguimiento."}), 400

    execute("""
        UPDATE pedidos SET estado='entregado', fecha_actualizado=CURRENT_TIMESTAMP
        WHERE id=?
    """, (pedido_id,))
    notificar_cliente_push(pedido_id, "entregado", pedido["nombre_local"])
    return jsonify({"ok": True, "msg": "¡Entrega confirmada!"})


# ── PÁGINAS LEGALES ───────────────────────────────────────────────────────────

@app.route("/terminos")
def terminos():
    return render_template("terminos.html")

@app.route("/privacidad")
def privacidad():
    return render_template("privacidad.html")


@app.route("/buscar")
def buscar_productos():
    q = request.args.get("q", "").strip()
    resultados = []
    if q and len(q) >= 2:
        resultados = query("""
            SELECT p.nombre AS prod_nombre, p.descripcion, p.precio, p.foto_url,
                   r.id AS restaurante_id, r.nombre_local, r.categoria,
                   r.logo_url, r.hace_envio, r.abierto
            FROM productos p
            JOIN restaurantes r ON r.id = p.restaurante_id
            WHERE p.disponible = 1
              AND r.estado = 'aprobado'
              AND (LOWER(p.nombre) LIKE %s OR LOWER(p.descripcion) LIKE %s)
            ORDER BY r.abierto DESC, p.nombre
            LIMIT 40
        """ if USE_POSTGRES else """
            SELECT p.nombre AS prod_nombre, p.descripcion, p.precio, p.foto_url,
                   r.id AS restaurante_id, r.nombre_local, r.categoria,
                   r.logo_url, r.hace_envio, r.abierto
            FROM productos p
            JOIN restaurantes r ON r.id = p.restaurante_id
            WHERE p.disponible = 1
              AND r.estado = 'aprobado'
              AND (LOWER(p.nombre) LIKE ? OR LOWER(p.descripcion) LIKE ?)
            ORDER BY r.abierto DESC, p.nombre
            LIMIT 40
        """, (f"%{q.lower()}%", f"%{q.lower()}%"))
    return render_template("buscar.html", q=q, resultados=resultados)


@app.route("/api/buscar-productos")
def api_buscar_productos():
    q = request.args.get("q", "").strip()
    if not q or len(q) < 2:
        return jsonify({"resultados": []})
    like = f"%{q.lower()}%"
    rows = query("""
        SELECT p.nombre AS prod_nombre, p.precio, p.foto_url,
               r.id AS restaurante_id, r.nombre_local, r.abierto
        FROM productos p
        JOIN restaurantes r ON r.id = p.restaurante_id
        WHERE p.disponible = 1 AND r.estado = 'aprobado'
          AND LOWER(p.nombre) LIKE %s
        ORDER BY r.abierto DESC, p.nombre LIMIT 8
    """ if USE_POSTGRES else """
        SELECT p.nombre AS prod_nombre, p.precio, p.foto_url,
               r.id AS restaurante_id, r.nombre_local, r.abierto
        FROM productos p
        JOIN restaurantes r ON r.id = p.restaurante_id
        WHERE p.disponible = 1 AND r.estado = 'aprobado'
          AND LOWER(p.nombre) LIKE ?
        ORDER BY r.abierto DESC, p.nombre LIMIT 8
    """, (like,))
    return jsonify({"resultados": [dict(r) for r in rows]})


@app.route("/mi-local/foto/logo/borrar", methods=["POST"])
@login_required
@rol_required("restaurante")
def borrar_logo():
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        return redirect(url_for("restaurante_panel"))
    execute("UPDATE restaurantes SET logo_url=NULL WHERE id=?", (restaurante["id"],))
    flash("Logo eliminado.", "success")
    return redirect(url_for("restaurante_panel") + "#sec-fotos")

@app.route("/mi-local/foto/banner/borrar", methods=["POST"])
@login_required
@rol_required("restaurante")
def borrar_banner():
    restaurante = get_restaurante_aprobado()
    if not restaurante:
        return redirect(url_for("restaurante_panel"))
    execute("UPDATE restaurantes SET banner_url=NULL WHERE id=?", (restaurante["id"],))
    flash("Banner eliminado.", "success")
    return redirect(url_for("restaurante_panel") + "#sec-fotos")


@app.route("/mi-cuenta/baja", methods=["GET", "POST"])
@login_required
def darse_de_baja():
    if request.method == "POST":
        from werkzeug.security import check_password_hash
        password = request.form.get("password", "")
        usuario  = query("SELECT * FROM usuarios WHERE id=?", (session["user_id"],), one=True)
        if not check_password_hash(usuario["password_hash"], password):
            flash("Contraseña incorrecta.", "danger")
            return redirect(url_for("darse_de_baja"))
        execute("UPDATE usuarios SET activo=0 WHERE id=?", (session["user_id"],))
        session.clear()
        flash("Tu cuenta fue desactivada. Lamentamos verte ir.", "info")
        return redirect(url_for("home"))
    return render_template("darse_de_baja.html")


@app.route("/mi-local/baja", methods=["GET", "POST"])
@login_required
@rol_required("restaurante")
def restaurante_darse_de_baja():
    if request.method == "POST":
        from werkzeug.security import check_password_hash
        password = request.form.get("password", "")
        usuario  = query("SELECT * FROM usuarios WHERE id=?", (session["user_id"],), one=True)
        if not check_password_hash(usuario["password_hash"], password):
            flash("Contraseña incorrecta.", "danger")
            return redirect(url_for("restaurante_darse_de_baja"))
        execute("UPDATE usuarios SET activo=0 WHERE id=?", (session["user_id"],))
        execute("UPDATE restaurantes SET estado='suspendido' WHERE usuario_id=?", (session["user_id"],))
        session.clear()
        flash("Tu local fue dado de baja. Podés volver cuando quieras registrándote de nuevo.", "info")
        return redirect(url_for("home"))
    return render_template("restaurante_baja.html")

# ── SETUP INICIAL ─────────────────────────────────────────────────────────────

@app.route("/setup/<clave_secreta>")
def setup_admin(clave_secreta):
    CLAVE    = os.environ.get("SETUP_KEY", "")
    password = os.environ.get("ADMIN_PASSWORD", "")
    if not CLAVE or not password:
        # Sin SETUP_KEY / ADMIN_PASSWORD configuradas en el entorno, esta
        # ruta queda deshabilitada — no hay fallback hardcodeado que un
        # atacante pueda adivinar leyendo el código fuente.
        return "Ruta no configurada. Definí SETUP_KEY y ADMIN_PASSWORD como variables de entorno.", 503
    if clave_secreta != CLAVE:
        return "No autorizado", 403

    admin_existe = query("SELECT id FROM usuarios WHERE rol='admin'", one=True)
    if admin_existe:
        return "Ya existe un administrador. Ruta desactivada.", 200

    from werkzeug.security import generate_password_hash

    execute("""
        INSERT INTO usuarios (nombre, apellido, email, telefono, password_hash, rol)
        VALUES (?, ?, ?, ?, ?, ?)
    """, ("Cristian", "Ojeda", "admin@pediaca.ar", "3417523674",
          generate_password_hash(password), "admin"))

    return f"""
    <html><head><meta charset="UTF-8"></head>
    <body style="font-family:sans-serif;padding:40px;text-align:center;max-width:500px;margin:0 auto;">
        <h2>✅ Admin creado</h2>
        <p><strong>Email:</strong> admin@pediaca.ar</p>
        <p><strong>Password:</strong> {password}</p>
        <p style="color:#e74c3c;font-weight:bold;margin-top:20px;">
            ⚠️ Anotá la password y cambiala desde tu perfil.
        </p>
        <a href="/login" style="display:inline-block;margin-top:20px;background:#F39C12;color:#fff;padding:14px 32px;border-radius:99px;text-decoration:none;font-weight:800;font-size:1rem;">
            Ir al login →
        </a>
    </body></html>
    """, 200


# ── CONFIGURACION DEL SITIO (BANNER HERO) ─────────────────────────────────

def get_config(clave):
    try:
        config = query("SELECT valor, tipo FROM configuraciones WHERE clave = ?", (clave,), one=True)
        if not config:
            return None
        if config['tipo'] == 'number':
            return int(config['valor']) if config['valor'] else 0
        return config['valor']
    except Exception as e:
        print(f"⚠️ Error en get_config('{clave}'): {type(e).__name__}: {e}")
        return None

def set_config(clave, valor, tipo='text'):
    try:
        sql = "INSERT INTO configuraciones (clave, valor, tipo, actualizado) VALUES (?, ?, ?, CURRENT_TIMESTAMP) ON CONFLICT(clave) DO UPDATE SET valor = ?, actualizado = CURRENT_TIMESTAMP"
        execute(sql, (clave, valor, tipo, valor))
    except Exception as e:
        print(f"⚠️ Error en set_config('{clave}'): {type(e).__name__}: {e}")

@app.context_processor
def inject_config():
    from functools import lru_cache
    @lru_cache(maxsize=50)
    def get_cached_config(clave):
        return get_config(clave)
    return dict(get_config=get_cached_config)


@app.context_processor
def inject_anuncios_lateral():
    """Banners de publicidad a los costados de la página (fixed, solo en
    pantallas anchas). Toggle independiente del carrusel del home — se
    inyecta acá (context_processor, no por ruta) para que aparezca en
    cualquier página sin tener que tocar cada vista. Reutiliza la tabla
    auspiciantes con posicion='header' (reutilizado como "lateral" para no
    tener que migrar el CHECK de la columna en la base ya en producción)."""
    activo = get_config("anuncios_lateral_activo") == "true"
    anuncios = []
    if activo:
        anuncios = query("""
            SELECT * FROM auspiciantes
            WHERE activo = 1 AND posicion = 'header'
              AND (fecha_inicio IS NULL OR fecha_inicio <= CURRENT_DATE)
              AND (fecha_fin   IS NULL OR fecha_fin   >= CURRENT_DATE)
        """)
    return dict(anuncios_lateral_activo=activo, anuncios_lateral=anuncios)


@app.route("/admin/configuracion", methods=["GET", "POST"])
@login_required
@rol_required("admin")
def admin_configuracion():
    mensaje = None
    
    if request.method == "POST":
        accion = request.form.get("accion", "")

        if accion == "borrar_imagen":
            set_config("hero_imagen_url", "")
            set_config("hero_activo", "false")
            mensaje = ("success", "✅ Imagen eliminada.")

        elif accion == "agregar_ciudad":
            nombre = request.form.get("nombre_ciudad", "").strip()
            if not nombre:
                mensaje = ("danger", "❌ Ingresá un nombre de ciudad.")
            else:
                existe = query("SELECT id FROM ciudades WHERE nombre = ?", (nombre,), one=True)
                if existe:
                    mensaje = ("danger", f"❌ '{nombre}' ya existe.")
                else:
                    fila = query("SELECT COALESCE(MAX(orden), 0) AS m FROM ciudades", one=True)
                    execute("INSERT INTO ciudades (nombre, activa, orden) VALUES (?, 1, ?)",
                            (nombre, (fila["m"] or 0) + 1))
                    mensaje = ("success", f"✅ '{nombre}' agregada y activada.")

        elif accion == "toggle_ciudad":
            try:
                ciudad_id = int(request.form.get("ciudad_id", 0))
            except ValueError:
                ciudad_id = 0
            fila = query("SELECT nombre, activa FROM ciudades WHERE id = ?", (ciudad_id,), one=True)
            if fila:
                execute("UPDATE ciudades SET activa = ? WHERE id = ?",
                        (0 if fila["activa"] else 1, ciudad_id))
                mensaje = ("success", "✅ Ciudad actualizada.")

        elif accion == "eliminar_ciudad":
            try:
                ciudad_id = int(request.form.get("ciudad_id", 0))
            except ValueError:
                ciudad_id = 0
            fila = query("SELECT nombre FROM ciudades WHERE id = ?", (ciudad_id,), one=True)
            if fila and fila["nombre"] == CIUDAD_DEFAULT:
                mensaje = ("danger", "❌ No podés eliminar la ciudad por defecto.")
            elif fila:
                execute("DELETE FROM ciudades WHERE id = ?", (ciudad_id,))
                mensaje = ("success", f"✅ '{fila['nombre']}' eliminada.")

        elif accion == "agregar_auspiciante":
            nombre_aus  = request.form.get("nombre_auspiciante", "").strip()
            url_destino = request.form.get("url_destino", "").strip()
            posicion    = request.form.get("posicion", "home")
            if posicion not in ("header", "home", "listado"):
                posicion = "home"
            if not nombre_aus:
                mensaje = ("danger", "❌ Ingresá un nombre.")
            else:
                logo_url = ""
                archivo = request.files.get("logo_auspiciante")
                if archivo and archivo.filename:
                    if allowed_file(archivo.filename):
                        ruta = guardar_imagen(archivo, "auspiciantes")
                        if ruta:
                            logo_url = url_imagen(ruta)
                    else:
                        mensaje = ("danger", "❌ Formato de imagen no soportado.")
                if not mensaje:
                    execute("""
                        INSERT INTO auspiciantes (nombre, logo_url, url_destino, activo, posicion)
                        VALUES (?, ?, ?, 1, ?)
                    """, (nombre_aus, logo_url, url_destino, posicion))
                    mensaje = ("success", f"✅ '{nombre_aus}' agregado.")

        elif accion == "toggle_auspiciante":
            try:
                aus_id = int(request.form.get("auspiciante_id", 0))
            except ValueError:
                aus_id = 0
            fila = query("SELECT activo FROM auspiciantes WHERE id = ?", (aus_id,), one=True)
            if fila:
                execute("UPDATE auspiciantes SET activo = ? WHERE id = ?",
                        (0 if fila["activo"] else 1, aus_id))
                mensaje = ("success", "✅ Auspiciante actualizado.")

        elif accion == "eliminar_auspiciante":
            try:
                aus_id = int(request.form.get("auspiciante_id", 0))
            except ValueError:
                aus_id = 0
            fila = query("SELECT nombre FROM auspiciantes WHERE id = ?", (aus_id,), one=True)
            if fila:
                execute("DELETE FROM auspiciantes WHERE id = ?", (aus_id,))
                mensaje = ("success", f"✅ '{fila['nombre']}' eliminado.")

        elif accion == "toggle_anuncios":
            actual = get_config("anuncios_activo") or "false"
            set_config("anuncios_activo", "false" if actual == "true" else "true")
            mensaje = ("success", "✅ Carrusel de publicidad (home) actualizado.")

        elif accion == "toggle_anuncios_lateral":
            actual = get_config("anuncios_lateral_activo") or "false"
            set_config("anuncios_lateral_activo", "false" if actual == "true" else "true")
            mensaje = ("success", "✅ Banners laterales actualizados.")

        else:
            hero_activo = "true" if request.form.get("hero_activo") == "true" else "false"

            archivo = request.files.get("hero_imagen")
            if archivo and archivo.filename:
                if allowed_file(archivo.filename):
                    ruta = guardar_imagen(archivo, "hero")
                    if ruta:
                        set_config("hero_imagen_url", url_imagen(ruta))
                        mensaje = ("success", "✅ Imagen subida.")
                    else:
                        mensaje = ("danger", "❌ No se pudo subir la imagen.")
                else:
                    mensaje = ("danger", "❌ Formato no soportado.")

            if hero_activo == "true" and not (get_config("hero_imagen_url") or (archivo and archivo.filename)):
                hero_activo = "false"
                mensaje = ("danger", "❌ Subí una imagen para poder activar el banner.")

            set_config("hero_activo", hero_activo)

            if not mensaje:
                mensaje = ("success", "✅ Configuración guardada.")

    config = {
        "hero_activo": get_config("hero_activo") or "false",
        "hero_imagen_url": get_config("hero_imagen_url") or "",
        "anuncios_activo": get_config("anuncios_activo") or "false",
        "anuncios_lateral_activo": get_config("anuncios_lateral_activo") or "false",
    }

    auspiciantes = query("SELECT * FROM auspiciantes ORDER BY activo DESC, nombre")

    return render_template("admin_configuracion.html", config=config, mensaje=mensaje,
                           ciudades=get_ciudades(), ciudad_default=CIUDAD_DEFAULT,
                           auspiciantes=auspiciantes)


# ── MAIN ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if not os.path.exists(DB_PATH):
        print(f"⚠️  No existe '{DB_PATH}'. Ejecutá primero: python init_db.py")
    else:
        app.run(debug=True, port=5000)